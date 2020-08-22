"""
Created on Jun 10, 2018
@author: Souvik
@Program Function: Adjust NSE EOD Data
"""

import os
from . import dates, utils
import pandas as pd
import numpy as np
import sqlite3
from sqlalchemy import create_engine
from collections import OrderedDict
import openpyxl as xl


class DataDB:
    """ Historical Bhavcopy data"""

    # constants

    YEARS = ['1995', '1996', '1997', '1998', '1999', '2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007',
             '2008', '2009', '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019', '2020']

    MOD_PATH = 'nse_eod_modifiers'
    INC_EXC_PATH = 'inc_exc_files'
    SYMBOL_CHANGE_FILE = '{}/symbol_change.csv'.format(MOD_PATH)
    SYMBOL_CHANGE_NEW_FILE = '{}/symbol_change_new.csv'.format(MOD_PATH)
    MULTIPLIERS_FILE = '{}/multipliers.csv'.format(MOD_PATH)
    MULTIPLIER_SKIPS_FILE = '{}/multipliers_skips.csv'.format(MOD_PATH)
    INDEX_CHANGE_DUMP_XL = '{}/IndexInclExcl.xlsx'.format(MOD_PATH)
    INDEX_CHANGE_DUMP_CSV = '{}/IndexInclExcl.csv'.format(MOD_PATH)
    SYMBOL_MAPPING_FILE = '{}/symbol_mapping.csv'.format(MOD_PATH)
    INDEX_CHANGE_MOD_CSV = '{}/IndexInclExclMod.csv'.format(MOD_PATH)
    INDEX_CHANGE_MANUAL_CSV = '{}/IndexInclExclManual.csv'.format(MOD_PATH)
    INDEX_COMPONENTS_CURR = '{}/index_components_curr.csv'.format(MOD_PATH)
    SYMBOL_CHECK_REPORT = '{}/symbol_check_report.csv'.format(MOD_PATH)
    SYMBOL_CHANGE_DUPLICATES_FILE = '{}/symbol_change_duplicates.csv'.format(MOD_PATH)
    UNVERIFIED_SKIPPED_RECORDS_FILE = '{}/unverified_skipped_records.csv'.format(MOD_PATH)
    UNVERIFIED_SELECTED_RECORDS_FILE = '{}/unverified_selected_records.csv'.format(MOD_PATH)
    REPLACEBLE_SELECTED_RECORDS = '{}/replaceble_selected_records.csv'.format(MOD_PATH)
    REPLACEBLE_SKIPPED_RECORDS = '{}/replaceble_skipped_records.csv'.format(MOD_PATH)
    DUMP_REPORT = '{}/dump_report.csv'.format(INC_EXC_PATH)
    MOD_DUMP_REPORT = '{}/mod_dump_report.csv'.format(INC_EXC_PATH)

    NIFTY_INDEX_NAME_CHANGES = {
        'Nifty Free Float Smallcap 100': 'NIFTY Smallcap 100',
        'Nifty Free Float Midcap 100': 'NIFTY Midcap 100'
    }

    NIFTY_INDICES_AMI = {
        'NIFTY LargeMidcap 250': 'NLMC250',
        'NIFTY Midcap 100': 'NMC100',
        'NIFTY Smallcap 100': 'NSC100',
        'Nifty 100': 'N100',
        'Nifty 200': 'N200',
        'Nifty 50': 'N50',
        'Nifty 500': 'N500',
        'Nifty Alpha 50': 'NALPHA50',
        'Nifty Auto': 'NAUTO',
        'Nifty Bank': 'NBANK',
        'Nifty Commodities': 'NCOMM',
        'Nifty Dividend Opportunities 50': 'NDIVO50',
        'Nifty Energy': 'NENERGY',
        'Nifty FMCG': 'NFMCG',
        'Nifty Financial Services': 'NFINS',
        'Nifty Growth Sectors 15': 'NGROWTH15',
        'Nifty High Beta 50': 'NHBETA50',
        'Nifty IT': 'NIT',
        'Nifty India Consumption': 'NINDCONS',
        'Nifty Infrastructure': 'NINFRA',
        'Nifty Low Volatility 50': 'NLVOL50',
        'Nifty MNC': 'NMNC',
        'Nifty Media': 'NMEDIA',
        'Nifty Metal': 'NMETAL',
        'Nifty Midcap 50': 'NMC50',
        'Nifty Midcap Liquid 15': 'NMCLIQ15',
        'Nifty Next 50': 'NNXT50',
        'Nifty PSE': 'NPSE',
        'Nifty PSU Bank': 'NPSUBANK',
        'Nifty Pharma': 'NPHARMA',
        'Nifty Realty': 'NREALTY',
        'Nifty Services Sector': 'NSERV',
        'Nifty100 Liquid 15': 'N100LIQ15',
        'Nifty50 Value 20': 'N50VAL20',
    }

    def __init__(self, path, db, years=None):

        if years is not None:
            self.YEARS = years

        os.chdir(path)
        
        print('Opening Bhavcopy database {}...'.format(db))
        self.conn = sqlite3.connect(db)
        self.engine = create_engine('sqlite:///{}'.format(db))

    def __del__(self):

        print('Closing DB connection..')
        self.conn.close()

    def truncate_table(self, table, msg=False):

        c = self.conn.cursor()
        c.execute('''DELETE FROM {}'''.format(table))
        self.conn.commit()
        c.close()

        if msg is True:
            print('Truncated table {}'.format(table))

    def load_multipliers(self, type='append'):
        """
        Load tblMultipliers table from csv
        :param type: 'append' or 'replace'
        :return: Nothing
        """

        multipliers = pd.read_csv(self.MULTIPLIERS_FILE)
        read_count = len(multipliers.index)
        print('Initiating loading of {} records into tblMultipliers'.format(read_count))

        if type == 'append':
            curr_multipliers = pd.read_sql_query('SELECT Symbol, Date, Multiplier FROM tblMultipliers', self.conn)
            print('{} records currently in tblMultipliers'.format(len(curr_multipliers.index)))
            multipliers = pd.concat([curr_multipliers, multipliers], axis=0)

        self.truncate_table('tblMultipliers', True)

        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        for symbol in symbol_changes.keys():
            multipliers['Symbol'] = np.where(multipliers.Symbol == symbol, symbol_changes[symbol], multipliers.Symbol)

        symbol_range_records = pd.read_sql('SELECT Symbol, StartDate FROM tblSymbolRange', self.conn)
        symbol_start_dates = dict(zip(symbol_range_records.Symbol, symbol_range_records.StartDate))

        c = self.conn.cursor()
        df_skips = pd.DataFrame()
        for symbol in multipliers['Symbol'].unique():
            print('processing', symbol)
            if symbol in symbol_start_dates:
                df = multipliers[(multipliers.Symbol == symbol) &
                             (multipliers.Date >= int(symbol_start_dates[symbol]))].copy()
                df['ResultantMultiplier'] = df['Multiplier'].cumprod()
                insert_rows = df.values.tolist()
                c.executemany('INSERT INTO tblMultipliers VALUES (?, ?, ?, ?)', insert_rows)
                self.conn.commit()
                df_skip = multipliers[(multipliers.Symbol == symbol) &
                                  (multipliers.Date < int(symbol_start_dates[symbol]))].copy()
                df_skips = pd.concat([df_skips, df_skip], axis=0)
            else:
                df_skip = multipliers[multipliers.Symbol == symbol].copy()
                df_skips = pd.concat([df_skips, df_skip], axis=0)

        c.close()

        df_skips.to_csv(self.MULTIPLIER_SKIPS_FILE, sep=',', index=False)

    def insert_records(self, df, table_name):
        """
        Insert passed records into table_name
        """

        df.to_sql(table_name, self.engine, index=False, if_exists='append')

    def load_tbldumps_from_csv(self, csv_path, start_year='1995'):
        """
        Load table with historical data
        :param start_year: start year from which data files need to be considered
        :param csv_path: path of historical data in csv format
        :return: None
        """

        years_to_load = [year for year in self.YEARS if year >= start_year]

        c = self.conn.cursor()

        print('Truncating tables')
        for year in years_to_load:
            self.truncate_table('tblDump{}'.format(year))

        csv_files = [f for f in os.listdir(csv_path) if f.endswith('.txt') and f[0:4] >= start_year]
        csv_files.sort()

        print('Initiating loading of {} files'.format(len(csv_files)))

        col_names = ['Symbol', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume']
        year = '1900'
        skipped_records = pd.DataFrame()
        duplicate_records = pd.DataFrame()
        read_count, write_count, skipped_count = 0, 0, 0
        for file in csv_files:
            if year != file[0:4]:
                if year != '1900': # Write records to yearly tblDumps
                    year_unique = year_records.drop_duplicates(['Symbol', 'Date'], keep='first')
                    year_duplicate = year_records[year_records.duplicated(['Symbol', 'Date'], keep=False)]
                    year_skipped = year_records[year_records.duplicated(['Symbol', 'Date'], keep='first')]
                    self.insert_records(year_unique, table_name='tblDump{}'.format(year))
                    write_count = write_count + len(year_unique.index)
                    skipped_records = pd.concat([skipped_records, year_skipped], axis=0)
                    duplicate_records = pd.concat([duplicate_records, year_duplicate], axis=0)
                    print('{}: inserted {} records, duplicate {} records, skipped {} records'.format(
                        year, len(year_unique.index), len(year_duplicate.index), len(year_skipped.index)))

                year = file[0:4]
                print('reading records for year {}'.format(year))
                year_records = pd.DataFrame()

            df = pd.read_csv(csv_path + file, names=col_names, header=None)
            year_records = pd.concat([year_records, df], axis=0)
            read_count = read_count + len(df.index)

        # Write records for final year
        year_unique = year_records.drop_duplicates(['Symbol', 'Date'], keep='first')
        year_duplicate = year_records[year_records.duplicated(['Symbol', 'Date'], keep=False)]
        year_skipped = year_records[year_records.duplicated(['Symbol', 'Date'], keep='first')]
        self.insert_records(year_unique, table_name='tblDump{}'.format(year))
        write_count = write_count + len(year_unique.index)
        skipped_records = pd.concat([skipped_records, year_skipped], axis=0)
        duplicate_records = pd.concat([duplicate_records, year_duplicate], axis=0)
        print('{}: inserted {} records, duplicate {} records, skipped {} records'.format(year, len(year_unique.index),
                                                                                         len(year_skipped.index),
                                                                                         len(year_duplicate.index)))

        c.close()
        self.insert_records(skipped_records, table_name='tblSkipped')
        self.insert_records(duplicate_records, table_name='tblDuplicates')

        print('{} files processed, {} records read, {} records inserted, {} records duplicate, {} records skipped'.format(
            len(csv_files), read_count, write_count, len(duplicate_records.index), len(skipped_records.index)))

    def append_tbldumps_from_csv(self, csv_path, start_date):
        """
        Append tblDump tables with historical data for last year records only
        :param start_date: start date in YYYY-MM-DD from which data files need to be considered
        :param csv_path: path of historical data in csv format
        :return: None
        """

        start_date = dates.yyyymmdd_to_yyyy_mm_dd(start_date)
        year = start_date[0:4]

        c = self.conn.cursor()

        year_qry = '''SELECT * FROM tblDump{} ORDER BY Symbol, Date ASC'''.format(year)
        year_records = pd.read_sql_query(year_qry, self.conn)

        self.truncate_table('tblDump{}'.format(year))

        csv_files = [f for f in os.listdir(csv_path) if f.endswith('.txt') and f[0:10] >= start_date]
        csv_files.sort()

        print('Initiating loading of {} files'.format(len(csv_files)))

        col_names = ['Symbol', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume']
        skipped_records = pd.DataFrame()
        duplicate_records = pd.DataFrame()
        read_count, write_count, skipped_count = 0, 0, 0
        for file in csv_files:
            df = pd.read_csv(csv_path + file, names=col_names, header=None)
            year_records = pd.concat([year_records, df], axis=0)
            read_count = read_count + len(df.index)

        # Write records for final year
        year_unique = year_records.drop_duplicates(['Symbol', 'Date'], keep='first')
        year_duplicate = year_records[year_records.duplicated(['Symbol', 'Date'], keep=False)]
        year_skipped = year_records[year_records.duplicated(['Symbol', 'Date'], keep='first')]
        self.insert_records(year_unique, table_name='tblDump{}'.format(year))
        write_count = write_count + len(year_unique.index)
        skipped_records = pd.concat([skipped_records, year_skipped], axis=0)
        duplicate_records = pd.concat([duplicate_records, year_duplicate], axis=0)
        print('{}: inserted {} records, duplicate {} records, skipped {} records'.format(year, len(year_unique.index),
                                                                                         len(year_skipped.index),
                                                                                         len(year_duplicate.index)))

        c.close()

        self.insert_records(skipped_records, table_name='tblSkipped')
        self.insert_records(duplicate_records, table_name='tblDuplicates')

        print('{} files processed, {} records read, {} records inserted, {} records duplicate, {} records skipped'.format(
            len(csv_files), read_count, write_count, len(duplicate_records.index), len(skipped_records.index)))

    def fetch_records(self, tables, symbols, start_date='19000101', end_date='21001231'):
        """
        Fetch records from tblDumps or tblModDumps for given parameters
        :param tables: 'tblDump' or 'tblModDump'
        :param symbols: [symbol1, symbol2, ...]
        :param start_date: 'YYYYMMDD'
        :param end_date: 'YYYYMMDD'
        :return: pandas dataframe with selected records
        """

        start_date, end_date = str(start_date), str(end_date)

        years = [year for year in self.YEARS if start_date[0:4] <= year <= end_date[0:4]]
        symbols_tuple = "('{}')".format(symbols[0]) if len(symbols) == 1 else str(tuple(symbols))

        if len(years) == 1:
            qry = """SELECT * FROM {}{} WHERE Symbol IN {} AND Date BETWEEN '{}' AND '{}'
                      ORDER BY Symbol ASC, Date ASC""".format(tables, years[0], symbols_tuple, start_date, end_date)
        else:
            qry_strings = []
            for i in range(0, len(years)):
                if i == 0:
                    qry_str = "SELECT * FROM {}{} WHERE Symbol IN {} AND Date >= '{}'".format(tables, years[i],
                                                                                              symbols_tuple, start_date)
                elif i == len(years) - 1:
                    qry_str = "SELECT * FROM {}{} WHERE Symbol IN {} AND Date <= '{}'".format(tables, years[i],
                                                                                              symbols_tuple, end_date)
                else:
                    qry_str = 'SELECT * FROM {}{} WHERE Symbol IN {}'.format(tables, years[i], symbols_tuple)

                qry_strings.append(qry_str)

            qry = '{} ORDER BY Symbol ASC, Date ASC'.format(' UNION '.join(qry_strings))

        return pd.read_sql_query(qry, self.conn)

    def int(self, val):
        try:
            return int(val)
        except ValueError:
            return 0

    def error(self, record, prev_record, next_record):
        """
        Calculate squared diff between prices
        :return: {'price_error': sum of squared diff for OHLC between mean of prev day and next day with curr day,
                'volume_error': squared diff for volume between mean of prev day and next day with curr day}
        """

        p_open, n_open, r_open = self.int(prev_record['Open'].iloc[0]), self.int(next_record['Open'].iloc[0]), \
                                 self.int(record['Open'].iloc[0])
        p_high, n_high, r_high = self.int(prev_record['High'].iloc[0]), self.int(next_record['High'].iloc[0]), \
                                 self.int(record['High'].iloc[0])
        p_low, n_low, r_low = self.int(prev_record['Low'].iloc[0]), self.int(next_record['Low'].iloc[0]), \
                                 self.int(record['Low'].iloc[0])
        p_close, n_close, r_close = self.int(prev_record['Close'].iloc[0]), self.int(next_record['Close'].iloc[0]), \
                              self.int(record['Close'].iloc[0])
        p_volume, n_volume, r_volume = self.int(prev_record['Volume'].iloc[0]), self.int(next_record['Volume'].iloc[0]), \
                                    self.int(record['Volume'].iloc[0])

        open_e = (p_open + n_open) / 2 - r_open
        high_e = (p_high + n_high) / 2 - r_high
        low_e = (p_low + n_low) / 2 - r_low
        close_e = (p_close + n_close) / 2 - r_close
        volume_e = (p_volume + n_volume) / 2 - r_volume

        return {'price_error': open_e ** 2 + high_e ** 2 + low_e ** 2 + close_e ** 2,
                'volume_error': volume_e ** 2}
        
    def calculate_skipped_record_errors(self, start_date='19000101'):
        """
        Assess whether selected records can be replaced by skipped records by calculating squared diffs with selected
        records price & volume with mean of prev and next day vis-a-vis that of skipped records
        :param start_date: YYYYMMDD for start date
        :return: Create csv file with replace recommendation
        """

        print('checking start date {}'.format(start_date))
        skipped_qry = 'SELECT DISTINCT * FROM tblSkipped WHERE Date >= "{}"'.format(start_date)
        skipped_records = pd.read_sql_query(skipped_qry, self.conn)

        unverified_skipped_records = pd.DataFrame()
        unverified_selected_records = pd.DataFrame()
        replaceable_selected_records = pd.DataFrame()
        replaceable_skipped_records = pd.DataFrame()

        for symbol in skipped_records['Symbol'].unique():
            print('checking for {}'.format(symbol))
            symbol_records = skipped_records[skipped_records.Symbol == symbol]
            for date in symbol_records['Date'].unique():
                date_records = symbol_records[symbol_records.Date == date]
                prev_sel_record_qry = '''SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date < "{}"
                                          ORDER BY Date DESC LIMIT 1'''.format(date[0:4], symbol, date)
                sel_record_qry = '''SELECT * FROM tblDump{} 
                                     WHERE Symbol = "{}" AND Date = {}'''.format(date[0:4], symbol, date)
                next_sel_record_qry = '''SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date > {}
                                          ORDER BY Date ASC LIMIT 1'''.format(date[0:4], symbol, date)

                prev_sel_record = pd.read_sql_query(prev_sel_record_qry, self.conn)
                sel_record = pd.read_sql_query(sel_record_qry, self.conn)
                next_sel_record = pd.read_sql_query(next_sel_record_qry, self.conn)

                if len(prev_sel_record.index) == 0 and len(next_sel_record.index) == 0:
                    for i in range(0, len(date_records.index)):
                        unverified_skipped_records = pd.concat([unverified_skipped_records,
                                                                date_records.iloc[i:i + 1, :]], axis=0)
                        unverified_selected_records = pd.concat([unverified_selected_records, sel_record], axis=0)
                else:
                    if len(prev_sel_record.index) == 0 and len(next_sel_record.index) > 0:
                        prev_sel_record = next_sel_record
                    elif len(prev_sel_record.index) > 0 and len(next_sel_record.index) == 0:
                        next_sel_record = prev_sel_record

                    sel_error = self.error(sel_record, prev_sel_record, next_sel_record)
                    skipped_error = []
                    for i in range(0, len(date_records.index)):
                        skipped_error.append(self.error(date_records.iloc[i:i+1, :], prev_sel_record, next_sel_record))
                        if skipped_error[i]['price_error'] < sel_error['price_error'] or \
                                skipped_error[i]['volume_error'] < sel_error['volume_error']:
                            replaceable_skipped_records = pd.concat([replaceable_skipped_records,
                                                                    date_records.iloc[i:i + 1, :]], axis=0)
                            replaceable_selected_records = pd.concat([replaceable_selected_records, sel_record], axis=0)

        unverified_skipped_records.to_csv(self.UNVERIFIED_SKIPPED_RECORDS_FILE, sep=',', index=False)
        unverified_selected_records.to_csv(self.UNVERIFIED_SELECTED_RECORDS_FILE, sep=',', index=False)
        replaceable_selected_records.to_csv(self.REPLACEBLE_SELECTED_RECORDS, sep=',', index=False)
        replaceable_skipped_records.to_csv(self.REPLACEBLE_SKIPPED_RECORDS, sep=',', index=False)

    def load_dump_replace_records(self):
        """
        Replace incorrect records in tblDumps with correct records which had been skipped earlier
        :return: None, update tblDumps with correct records
        """

        self.truncate_table('tblDumpReplace', True)

        c = self.conn.cursor()

        try:
            replaceable_selected_records = pd.read_csv(self.REPLACEBLE_SELECTED_RECORDS)
            replaceable_skipped_records = pd.read_csv(self.REPLACEBLE_SKIPPED_RECORDS)
        except pd.errors.EmptyDataError:
            print('No records to replace')
            return

        for i in range(0, len(replaceable_selected_records.index)):
            sel_record = replaceable_selected_records.iloc[i:i+1, :]
            skipped_record = replaceable_skipped_records.iloc[i:i + 1, :]
            insert_row = (skipped_record['Symbol'].iloc[0], str(skipped_record['Date'].iloc[0]),
                          skipped_record['Open'].iloc[0], skipped_record['High'].iloc[0], skipped_record['Low'].iloc[0],
                          skipped_record['Close'].iloc[0], int(skipped_record['Volume'].iloc[0]))

            if sel_record['Symbol'].iloc[0] == skipped_record['Symbol'].iloc[0] and \
                sel_record['Date'].iloc[0] == skipped_record['Date'].iloc[0]:
                try:
                    c.execute('''INSERT INTO tblDumpReplace VALUES (?,?,?,?,?,?,?)''', insert_row)
                    print(insert_row, "inserted")
                except sqlite3.IntegrityError:
                    print(insert_row, "insert failed")
            else:
                print(insert_row, "skipped")

        self.conn.commit()
        c.close()

    def save_symbols_date_range(self, append_from='1995', type='append'):
        """
        Save all symbols start and end dates in csv file
        :param append_from: start year for appending
        :return:
        """

        exist_symbols_range = pd.DataFrame()
        if type != 'append':
            exist_symbols_range = pd.read_sql_query('''SELECT * FROM tblSymbolRange''', self.conn)
        self.truncate_table('tblSymbolRange', True)

        years = [year for year in self.YEARS if year >= append_from]

        qry_strings_dump, qry_strings_mod = [], []
        for year in years:
            qry_strings_dump.append("SELECT Symbol, 'Dump' TableSource, MIN(Date) MinDate, MAX(Date) MaxDate "
                                    "FROM tblDump{} GROUP BY Symbol".format(year))
            qry_strings_mod.append("SELECT Symbol, 'ModDump' TableSource, MIN(Date) MinDate, MAX(Date) MaxDate "
                                   "FROM tblModDump{} GROUP BY Symbol".format(year))

        qry_dump = """SELECT Symbol, 'Dump' TableSource, MIN(MinDate) StartDate, MAX(MaxDate) EndDate FROM ({})
         GROUP BY Symbol""".format(' UNION '.join(qry_strings_dump))
        new_symbols_range_dump = pd.read_sql_query(qry_dump, self.conn)

        qry_mod = """SELECT Symbol, 'ModDump' TableSource, MIN(MinDate) StartDate, MAX(MaxDate) EndDate FROM ({})
         GROUP BY Symbol""".format(' UNION '.join(qry_strings_mod))
        new_symbols_range_mod = pd.read_sql_query(qry_mod, self.conn)

        new_symbols_range = pd.concat([new_symbols_range_dump, new_symbols_range_mod], axis=0)
        symbols_range = pd.concat([exist_symbols_range, new_symbols_range], axis=0)
        self.insert_records(symbols_range, 'tblSymbolRange')

        if len(new_symbols_range.index) > 0:
            print('Adjusting for existing records')
            qry = 'SELECT Symbol, TableSource, MIN(StartDate) StartDate, MAX(EndDate) EndDate FROM tblSymbolRange ' \
                  'GROUP BY Symbol, TableSource'
            final_recs = pd.read_sql_query(qry, self.conn)
            self.truncate_table('tblSymbolRange', True)
            self.insert_records(final_recs, 'tblSymbolRange')
        
        print('symbol update complete')

    def save_symbols_date_range_delta(self, start_date):
        """
        Update symbols start and end dates in tblSymbolRange
        :param start_date: start year for appending
        :return:
        """

        c = self.conn.cursor()

        qry = '''SELECT Symbol, Min(Date) MinDate, Max(Date) MaxDate, Source, TableSource, StartDate, EndDate
                    FROM (SELECT t1.Symbol, t1.Date, 'Dump' Source, t2.TableSource, t2.StartDate, t2.EndDate
                    FROM tblDump{year} t1
                    LEFT JOIN tblSymbolRange t2
                    ON t1.Symbol = t2.Symbol
                    AND TableSource <> 'ModDump'
                    WHERE t1.Date >= {start_date}
                    UNION
                SELECT t1.Symbol, t1.Date, 'ModDump' Source, t2.TableSource, t2.StartDate, t2.EndDate
                    FROM tblModDump{year} t1
                    LEFT JOIN tblSymbolRange t2
                    ON t1.Symbol = t2.Symbol
                    AND TableSource <> 'Dump'
                    WHERE t1.Date >= {start_date}
                    UNION
                SELECT t1.Symbol, t1.Date, 'Dump' Source, t2.TableSource, t2.StartDate, t2.EndDate
                    FROM tblDump{year} t1
                    LEFT JOIN tblSymbolRange t2
                    ON t1.Symbol = t2.Symbol
                    AND Source = t2.TableSource
                    WHERE t1.Date >= {start_date}
                    UNION
                SELECT t1.Symbol, t1.Date, 'ModDump' Source, t2.TableSource, t2.StartDate, t2.EndDate
                    FROM tblModDump{year} t1
                    LEFT JOIN tblSymbolRange t2
                    ON t1.Symbol = t2.Symbol
                    AND Source = t2.TableSource
                    WHERE t1.Date >= {start_date}) t1
                GROUP BY Symbol, Source, TableSource'''.format(start_date=start_date, year=start_date[0:4])

        df = pd.read_sql_query(qry, self.conn)

        for _, row in df.iterrows():   
            if row.TableSource is None: # Symbol not in tblSymbolRange; Insert Record
                c.execute('INSERT INTO tblSymbolRange VALUES ("{}", "{}", {}, {})'.format(
                    row.Symbol, row.Source, row.MinDate, row.MaxDate))  
            else:
                row.MinDate, row.MaxDate, row.StartDate, row.EndDate = int(row.MinDate), int(row.MaxDate), \
                    int(row.StartDate), int(row.EndDate)
                if row.MinDate > row.StartDate: # Symbol exists for earlier dates; No action
                    pass
                if row.MinDate <= row.StartDate: # Error: Symbol start date later than new dates; Investigate
                    print('MinDate <= StartDate - ', row.Symbol, row.MinDate, row.MaxDate, row.Source, row.TableSource, row.StartDate, row.EndDate)
                if row.MaxDate <= row.EndDate: # Error Symbol date range ends later than actual data; Investigate
                    print('MaxDate <= EndDate -', row.Symbol, row.MinDate, row.MaxDate, row.Source, row.TableSource, row.StartDate, row.EndDate)
                if row.MaxDate > row.EndDate: # New records with later dates; Update Record
                    c.execute('UPDATE tblSymbolRange SET EndDate = {} WHERE Symbol = "{}" AND TableSource = "{}"'.format(
                        row.MaxDate, row.Symbol, row.Source))

        self.conn.commit()
        c.close()
        
        print('symbol range update complete')

    def delta_date_check(self, start_date):
        """
        End in error if delta update date is older than latest available date
        """
        df = pd.read_sql_query("SELECT MAX(EndDate) MaxDate FROM tblSymbolRange", self.conn)

        assert int(df.MaxDate.iloc[0]) < int(start_date), 'Delta date {} must be greater than database date {}'.format(
            start_date, df.MaxDate.iloc[0])

    def load_modified_tbldumps(self, start_year='1995'):
        """
        Load tblModDumps from tblDumps with symbol changes and corrected records
        :param start_year: start year from which tables need to be updated
        :return: None
        """

        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        years_to_load = [year for year in self.YEARS if start_year <= year <= self.YEARS[-1]]

        c = self.conn.cursor()

        print('Truncating tables')
        for year in years_to_load:
            self.truncate_table('tblModDump{}'.format(year))

        df_duplicate = pd.DataFrame()
        for year in years_to_load:
            print('updating for year', year)
            df = pd.read_sql_query('''SELECT * FROM tblDump{}'''.format(year), self.conn)
            df_replace = pd.read_sql_query('''SELECT * FROM tblDumpReplace WHERE Date LIKE "{}%"'''.format(year),
                                           self.conn)
            for symbol in symbol_changes.keys():
                df['Symbol'] = np.where(df.Symbol == symbol, symbol_changes[symbol], df.Symbol)
                df_replace['Symbol'] = np.where(df_replace.Symbol == symbol, symbol_changes[symbol], df_replace.Symbol)
            df = pd.concat([df_replace, df], axis=0)
            df['AdjustedOpen'] = np.nan
            df['AdjustedHigh'] = np.nan
            df['AdjustedLow'] = np.nan
            df['AdjustedClose'] = np.nan
            df_unique = df.drop_duplicates(['Symbol', 'Date'], keep='first').copy()
            self.insert_records(df_unique, 'tblModDump{}'.format(year))
            df_duplicate = pd.concat([df_duplicate, df[df.duplicated(['Symbol', 'Date'], keep='first')]], axis=0)

        if len(df_duplicate.index) > 0:
            df_duplicate.to_csv(self.SYMBOL_CHANGE_DUPLICATES_FILE, sep=',', index=False)

        c.close()

    def append_modified_tbldumps(self, start_date):
        """
        Append tblModDumps from tblDumps with symbol changes and corrected records. Can be done only for last year
        :param start_date: start date from which tables need to be updated in YYYYMMDD
        :return: None
        """

        c = self.conn.cursor()

        year = start_date[0:4]
        year_qry = '''SELECT * FROM tblModDump{} ORDER BY Symbol, Date ASC'''.format(year)
        #print(year_qry)
        year_records = pd.read_sql_query(year_qry, self.conn)
        self.truncate_table('tblModDump{}'.format(year))

        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        df_duplicate = pd.DataFrame()

        df = pd.read_sql_query('''SELECT * FROM tblDump{} WHERE Date >= "{}"'''.format(year, start_date), self.conn)
        df_replace = pd.read_sql_query('''SELECT * FROM tblDumpReplace'''.format(year), self.conn)
        for symbol in symbol_changes.keys():
            df['Symbol'] = np.where(df.Symbol == symbol, symbol_changes[symbol], df.Symbol)
            df_replace['Symbol'] = np.where(df_replace.Symbol == symbol, symbol_changes[symbol], df_replace.Symbol)
        df = pd.concat([df_replace, df], axis=0)
        df['AdjustedOpen'] = np.nan
        df['AdjustedHigh'] = np.nan
        df['AdjustedLow'] = np.nan
        df['AdjustedClose'] = np.nan
        df = pd.concat([year_records, df], axis=0)
        df_unique = df.drop_duplicates(['Symbol', 'Date'], keep='first').copy()
        self.insert_records(df_unique, 'tblModDump{}'.format(year))
        df_duplicate = pd.concat([df_duplicate, df[df.duplicated(['Symbol', 'Date'], keep='first')]], axis=0)

        if len(df_duplicate.index) > 0:
            df_duplicate.to_csv(self.SYMBOL_CHANGE_DUPLICATES_FILE, sep=',', index=False)

        c.close()

    def update_records(self, df, tables='tblModDump'):
        """
        Update passed records into table_name
        """

        c = self.conn.cursor()

        symbol = df.iloc[0]['Symbol']
        start_date = df.iloc[0]['Date']
        years = [year for year in self.YEARS if year >= start_date[0:4]]

        for year in years:
            if year == start_date[0:4]:
                delete_qry = '''DELETE FROM {}{} WHERE Symbol = "{}" AND Date >= "{}"'''.format(tables, year, symbol,
                                                                                                start_date)
            else:
                delete_qry = '''DELETE FROM {}{} WHERE Symbol = "{}"'''.format(tables, year, symbol)
            c.execute(delete_qry)
            df_year = df[(df.Date >= '{}0101'.format(year)) & (df.Date <= '{}1231'.format(year))]
            c.executemany('INSERT INTO {}{} VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'.format(tables, year),
                          df_year.values.tolist())
            self.conn.commit()

    def update_adjusted_price(self, start_date='19000101'):
        """
        Update adjusted price based on multipliers
        :param start_date:
        :return:
        """

        multipliers_qry = '''SELECT * FROM tblMultipliers ORDER BY Symbol, Date'''
        multipliers = pd.read_sql_query(multipliers_qry, self.conn)

        c = self.conn.cursor()

        for symbol in multipliers['Symbol'].unique():
            symbol_multipliers = multipliers[multipliers.Symbol == symbol]
            fetch_start_date = max(start_date, symbol_multipliers.iloc[0]['Date'])
            symbol_records = self.fetch_records('tblModDump', [symbol], start_date=fetch_start_date)
            symbol_records_update = pd.DataFrame()
            for idx, row in symbol_multipliers.iterrows():
                if len(symbol_records_update) > 0:  # Discard records >= row[date]
                    symbol_records_update = symbol_records_update[symbol_records_update.Date < row['Date']]
                temp_records = symbol_records[symbol_records.Date >= row['Date']].copy()
                temp_records['AdjustedOpen'] = temp_records['Open'] * row['ResultantMultiplier']
                temp_records['AdjustedHigh'] = temp_records['High'] * row['ResultantMultiplier']
                temp_records['AdjustedLow'] = temp_records['Low'] * row['ResultantMultiplier']
                temp_records['AdjustedClose'] = temp_records['Close'] * row['ResultantMultiplier']
                symbol_records_update = pd.concat([symbol_records_update, temp_records], axis=0)
            if len(symbol_records_update) > 0:
                self.update_records(symbol_records_update)
                print('update complete for', symbol)
            else:
                print('update skipped for', symbol)

        c.close()

    def index_change_xl_to_csv(self):
        """
        Index Change excel file to modified csv, use carefully. Ensure manual checks. 
        Date column must be verified after run. Symbols not found should be updated after run.
        :return:
        """

        wb_idx = xl.load_workbook(self.INDEX_CHANGE_DUMP_XL)

        sheets = wb_idx.sheetnames

        cols = ['Index', 'Date', 'Symbol', 'ChangeType']
        records = []

        for sheet in sheets:
            sheet_name = wb_idx[sheet]
            for row in range(2, sheet_name.max_row + 1):
                records.append([sheet_name['A' + str(row)].value,
                                sheet_name['B' + str(row)].value,
                                sheet_name['C' + str(row)].value,
                                sheet_name['D' + str(row)].value])

        df = pd.DataFrame(records, columns=cols)
        df.to_csv(self.INDEX_CHANGE_DUMP_CSV, sep=',', index=False)

        # Create corrected file

        symbol_mapping_records = pd.read_csv(self.SYMBOL_MAPPING_FILE)
        symbol_mapping = dict(zip(symbol_mapping_records.Scrip, symbol_mapping_records.Symbol))

        all_scrips = set(df.Symbol)
        scrips_with_symbols = set(symbol_mapping_records.Scrip)

        for scrip in symbol_mapping.keys():
            df.loc[df.Symbol == scrip, 'Symbol'] = symbol_mapping[scrip]

        df.loc[df['ChangeType'].str.contains('Inclusion'), 'ChangeType'] = 'I'
        df.loc[df['ChangeType'].str.contains('Exclusion'), 'ChangeType'] = 'E'

        # Convert date to YYYYMMDD
        df.Date = [int(x.strftime('%Y%m%d')) if isinstance(x, dates.datetime) \
            else int(dates.dd_mm_yyyy_to_yyyymmdd(x)) for x in df.Date]

        # Convert to latest symbol
        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        for symbol in symbol_changes.keys():
            df['Symbol'] = np.where(df.Symbol == symbol, symbol_changes[symbol], df.Symbol)

        # Update latest index names
        for index in self.NIFTY_INDEX_NAME_CHANGES.keys():
            df['Index'] = np.where(df.Index == index, self.NIFTY_INDEX_NAME_CHANGES[index], df.Index)

        # Update from manual file
        manual = pd.read_csv(self.INDEX_CHANGE_MANUAL_CSV, encoding='ansi')

        for symbol in manual.Symbol.unique():
            symbol_dates = manual[manual.Symbol == symbol].Date.unique()
            correct_symbol = manual[manual.Symbol == symbol].CorrectSymbol.iloc[0]
            for date in symbol_dates:
                df['Symbol'] = np.where((df.Symbol == symbol) & (df.Date == date), correct_symbol, df.Symbol)

        # Write file
        
        df.to_csv(self.INDEX_CHANGE_MOD_CSV, sep=',', index=False)

        # Code to display scrips not converted to symbols
        print('Scrips could not be mapped to symbols:')
        for scrip in (all_scrips - scrips_with_symbols):
            print(scrip)

    def check_symbol_dates(self):
        """
        Validate index change file to see if symbols are present on those days
        :return:
        """

        print('checking symbol dates')

        df = pd.read_csv(self.INDEX_CHANGE_MOD_CSV)
        result = pd.DataFrame(columns=['ChangeType', 'Symbol', 'RangeStart', 'RangeEnd', 'Found/NotFound', 'SearchDate', \
            'Index', 'DiffBefore', 'DiffAfter'])

        for idx, row in df.iterrows():
            symbol_found = self.fetch_records('tblModDump', [row.Symbol], start_date=row.Date, end_date=row.Date)

            if not symbol_found.empty:        # Symbol found on date
                #print('Exact Match : ChangeType: {} - {} - {}'.format(row.ChangeType, row.Symbol, row.Date))
                pass
            else:                             # Symbol not found on date
                symbol_range = pd.read_sql("SELECT Symbol, CAST(StartDate AS INT) StartDate, CAST(EndDate AS INT) EndDate FROM tblSymbolRange "
                                            "WHERE Symbol = '{}' AND TableSource = 'ModDump'".format(row.Symbol), self.conn)
                if symbol_range.empty:
                    #print('ChangeType: {} - ,{}, : Range Start: ,XXXXXXXX, Range End: ,XXXXXXXX, : Not Found, {},{}'.format(
                    #    row.ChangeType, row.Symbol, row.Date, row.Index))
                    result = result.append([{'ChangeType': row.ChangeType, 'Symbol': row.Symbol, \
                        'RangeStart': 99990101, 'RangeEnd': 99990101, 'Found/NotFound': 'Not Found', \
                            'SearchDate': row.Date, 'Index': row.Index}])
                elif symbol_range.iloc[0].StartDate < row.Date < symbol_range.iloc[0].EndDate:
                    #print('ChangeType: {} - ,{}, : Range Start: ,{}, Range End: ,{}, : Found in Range'.format(
                    #    row.ChangeType, row.Symbol, symbol_range.iloc[0].StartDate, symbol_range.iloc[0].EndDate))
                    pass
                else:
                    #print('ChangeType: {} - ,{}, : Range Start: ,{}, Range End: ,{}, : Found out of Range ,{},{}'.format(
                    #    row.ChangeType, row.Symbol, symbol_range.iloc[0].StartDate, symbol_range.iloc[0].EndDate, row.Date, row.Index))
                    result = result.append([{'ChangeType': row.ChangeType, 'Symbol': row.Symbol, \
                        'RangeStart': symbol_range.iloc[0].StartDate, 'RangeEnd': symbol_range.iloc[0].EndDate, \
                            'Found/NotFound': 'Found out of Range', 'SearchDate': row.Date, 'Index': row.Index}])

        result.DiffBefore = result[['RangeStart', 'SearchDate']].apply(dates.datediff, axis=1) 
        result.DiffAfter = result[['SearchDate', 'RangeEnd']].apply(dates.datediff, axis=1)
        result.DiffBefore = np.where(result.DiffBefore < 0, 0, result.DiffBefore)
        result.DiffAfter = np.where(result.DiffAfter < 0, 0, result.DiffAfter)

        result.to_csv(self.SYMBOL_CHECK_REPORT, sep=',', index=False)

    def table_report(self):

        df_dump, df_mod_dump = pd.DataFrame(), pd.DataFrame()
        for year in self.YEARS:
            print('reading data for', year)
            dump_qry = """SELECT '{0}' Year, COUNT(*) Count, SUM(Open) OpenSum, SUM(High) HighSum, SUM(Low) LowSUm, SUM(Close) CloseSum, SUM(Volume) VolumeSum
                            FROM tblDump{0}""".format(year)
            mod_dump_qry = """SELECT '{0}' Year, COUNT(*) Count, SUM(Open) OpenSum, SUM(High) HighSum, SUM(Low) LowSUm, SUM(Close) CloseSum, SUM(Volume) VolumeSum,
                                     SUM(AdjustedOpen) AdjustedOpenSum, SUM(AdjustedHigh) AdjustedHighSum, SUM(AdjustedLow) AdjustedLowSum, SUM(AdjustedClose) AdjustedCloseSum
                                FROM tblModDump{0}""".format(year)
            df_dump = pd.concat([df_dump, pd.read_sql(dump_qry, self.conn)], axis=0)
            df_mod_dump = pd.concat([df_mod_dump, pd.read_sql(mod_dump_qry, self.conn)], axis=0)

        df_dump.index, df_mod_dump.index = df_dump.Year, df_mod_dump.Year
        df_dump.drop('Year', axis=1)
        df_mod_dump.drop('Year', axis=1)

        df_dump.to_csv(self.DUMP_REPORT, sep=',', index=False)
        df_mod_dump.to_csv(self.MOD_DUMP_REPORT, sep=',', index=False)

        print('printing Dump report...')
        print(df_dump)
        print('printing modDump report...')
        print(df_mod_dump)

        print('printing diff...')
        for i in df_dump.index:
            for col in df_dump.columns:
                diff = int(df_dump.loc[i][col]) - int(df_mod_dump.loc[i][col])
                if diff != 0:
                    print(col, diff)

    def load_historical_index_components(self):
        """
        Load tblHistIndex based on index change hist and current index components
        :return:
        """

        c = self.conn.cursor()

        self.truncate_table('tblHistIndex', True)

        idx_change_hist = pd.read_csv(self.INDEX_CHANGE_MOD_CSV)
        idx_components_curr = pd.read_csv(self.INDEX_COMPONENTS_CURR)

        for index in idx_components_curr['Index'].unique():
            print('Loading for', index)
            index_hist = idx_change_hist[idx_change_hist.Index == index].copy()
            change_dates = list(index_hist['Date'].unique())
            change_dates.sort(reverse=True)
            curr_symbol_recs = idx_components_curr[idx_components_curr.Index == index].copy()

            first = True
            insert_symbols = set()
            for change_date in change_dates:
                excluded_symbols = index_hist[(index_hist.Date == change_date) &
                                              (index_hist.ChangeType == 'E')]['Symbol'].tolist()
                included_symbols = index_hist[(index_hist.Date == change_date) &
                                              (index_hist.ChangeType == 'I')]['Symbol'].tolist()

                if first:
                    curr_symbols = curr_symbol_recs['Symbol'].tolist()
                    curr_dates = [change_date] * len(curr_symbols)
                    curr_index = [index] * len(curr_symbols)
                    curr_df = pd.DataFrame(np.column_stack([curr_index, curr_symbols, curr_dates]),
                                           columns=['Index', 'Symbol', 'Date'])
                    insert_rows = curr_df.values.tolist()
                    c.executemany('INSERT INTO tblHistIndex VALUES (?, ?, ?)', insert_rows)
                    insert_symbols = list(set(curr_symbols).union(set(excluded_symbols)) - set(included_symbols))
                    first = False
                else:
                    insert_symbols = list(set(insert_symbols).union(set(included_symbols)) - set(excluded_symbols))
                    insert_dates = [change_date] * len(insert_symbols)
                    insert_index = [index] * len(insert_symbols)
                    insert_df = pd.DataFrame(np.column_stack([insert_index, insert_symbols, insert_dates]),
                                             columns=['Index', 'Symbol', 'Date'])
                    insert_rows = insert_df.values.tolist()
                    c.executemany('INSERT INTO tblHistIndex VALUES (?, ?, ?)', insert_rows)
                    insert_symbols = list(set(insert_symbols).union(excluded_symbols) - set(included_symbols))

        self.conn.commit()
        c.close()

    def symbols_index_hist_files(self, indices='default', start_year='1995'):
        """
        Create yearly files for symbols inclusion/exclusion in indices
        :param indices:
        :param start_year:
        :param end_year:
        :return:
        """

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )

        years = [year for year in self.YEARS if start_year <= year <= self.YEARS[-1]]

        c = self.conn.cursor()

        print('Loading Index History...')
        qry = """SELECT t1. * 
                   FROM tblHistIndex t1
                   JOIN (SELECT IndexName, Min(Date) FirstDate
                           FROM (SELECT DISTINCT IndexName, Max(Date) Date
                                   FROM tblHistIndex
                                  WHERE Date < '{}0101'
                                  GROUP BY IndexName
                                  UNION
                                 SELECT DISTINCT IndexName, Min(Date) Date
                                   FROM tblHistIndex
                                  WHERE Date >= '{}0101'
                                  GROUP BY IndexName)
                          GROUP BY IndexName) t2
                  WHERE t1.IndexName = t2.IndexName
                    AND t1.Date >= t2.FirstDate
                    AND t2.IndexName in {}
                  ORDER BY IndexName ASC, Date ASC""".format(start_year, start_year, str(indices))

        index_hist = pd.read_sql(qry, self.conn)
        index_hist_first = index_hist.drop_duplicates(['IndexName'], keep='first')
        first_dates = dict(zip(index_hist_first['IndexName'], index_hist_first['Date']))
        prev_dates = first_dates

        for year in years:
            print('Processing {}...'.format(year))

            qry_yr = 'SELECT Symbol, Date FROM tblModDump{}'.format(year)
            yr_records = pd.read_sql(qry_yr, self.conn)
            yr_records['Date'] = pd.to_numeric(yr_records['Date'])
            for index in index_hist['IndexName'].unique():
                print('Processing {} {}...'.format(year, index))
                yr_records[index] = [0] * yr_records['Symbol'].size
                index_hist_curr = index_hist[(index_hist.IndexName == index) &
                                             (index_hist.Date >= prev_dates[index])].copy()
                in_scope_dates = [d for d in list(index_hist_curr['Date'].unique()) if int(d[0:4]) == int(year)]
                for date in in_scope_dates:
                    print('................ Processing {}... prev date {}'.format(date, prev_dates[index]))

                    index_hist_prev_date = index_hist_curr[index_hist_curr.Date == prev_dates[index]].copy()
                    for symbol in index_hist_prev_date['Symbol'].unique():
                        yr_records.loc[(yr_records.Symbol == symbol) & (yr_records.Date >= int(prev_dates[index])) &
                                       (yr_records.Date < int(date)), index] = 1
                    prev_dates[index] = date
                if prev_dates[index][0:4] != year:
                    continue

                print('................ Processing {}...'.format(prev_dates[index]))
                index_hist_prev_date = index_hist_curr[index_hist_curr.Date == prev_dates[index]].copy()
                for symbol in index_hist_prev_date['Symbol'].unique():
                    yr_records.loc[(yr_records.Symbol == symbol) &
                                   (yr_records.Date >= int(prev_dates[index])), index] = 1

            yr_records.to_csv('{}{}.csv'.format(self.INC_EXC_PATH, year), sep=',', index=False)

        c.close()

    def symbols_index_hist_files_delta(self, start_date, indices='default'):
        """
        Create current file for symbols inclusion/exclusion in indices
        :param indices:
        :param start_date:
        :return:
        """

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )

        c = self.conn.cursor()

        print('Loading Current Index Components...')

        idx_comp = pd.read_csv(self.INDEX_COMPONENTS_CURR)
        idx_comp['IncFlag'] = 1
        #index = idx_comp.Index + idx_comp.Symbol
        idx_comp = idx_comp.set_index('Symbol')

        print('Processing data from {}...'.format(start_date))

        qry = "SELECT Symbol, Date FROM tblModDump{} WHERE Date >= '{}'".format(start_date[0:4], start_date)
        records = pd.read_sql(qry, self.conn)
        records = records.set_index('Symbol')
        #records['Date'] = pd.to_numeric(records['Date'])
        records['Symbol'] = records.index
        records = records[['Symbol', 'Date']]
        for index in indices:
            print('Processing {}...'.format(index))
            idx_comp_index = idx_comp[idx_comp.Index == index].copy()
            records = records.assign(IncFlag=idx_comp_index.IncFlag).fillna(0)
            records = records.rename(index=str, columns={'IncFlag': index})



        records.to_csv('{}{}.csv'.format(self.INC_EXC_PATH, start_date), sep=',', index=False)

        c.close()

    def update_symbol_change(self):

        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_NEW_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        c = self.conn.cursor()

        for symbol in symbol_changes.keys():
            for year in self.YEARS:
                print('updating for', symbol, year)
                update_qry = "UPDATE tblModDump{} SET Symbol = '{}' WHERE Symbol = '{}'".format(
                    year, symbol_changes[symbol], symbol)
                c.execute(update_qry)
                self.conn.commit()

        c.close()

        self.save_symbols_date_range(type='refresh')

    def test_inc_exc_mod_list(self):

        idx_change_hist = pd.read_csv(self.INDEX_CHANGE_MOD_CSV)
        idx_components_curr = pd.read_csv(self.INDEX_COMPONENTS_CURR)

        idx_last_include = idx_change_hist.sort_values(by=['Symbol', 'Date'])
        idx_last_include = idx_last_include.drop_duplicates(['Index', 'Symbol'], keep='last')

        idx_last_include.to_csv('{}/IndexInclExclModLastInc.csv'.format(self.MOD_PATH), sep=',', index=False)

        for index in idx_last_include['Index'].unique():
            last_inc_curr_index = idx_last_include[(idx_last_include.Index == index) & \
                (idx_last_include.ChangeType == 'I')]['Symbol']
            idx_components_curr_index = idx_components_curr[idx_components_curr.Index == index]['Symbol']

            idx_components = set(idx_components_curr_index)
            last_includes = set(last_inc_curr_index)

            print('$$$$$$$', index, '$$$$$$ curr components', len(idx_components), ' last includes ', len(last_includes))
            print('####### last includes not in curr index #######')
            print(last_includes - idx_components)
            print('####### curr index components not in last includes #######')
            print(idx_components - last_includes)


    def compare_index_data(self, index, year_file, prev_date, date, index_hist, index_year_data):
        """Test symbols selected for certain dates in indices """

        print('{}........................ {} Processing {} - {}...'.format(index, year_file, prev_date, date))

        symbols_index_hist = index_hist[index_hist.Date == prev_date]['Symbol'].unique()
        symbols_index_data = index_year_data[(index_year_data.Date >= int(prev_date)) &
                                             (index_year_data.Date < int(date))]['Symbol'].unique()

        symbols_hist = set(symbols_index_hist)
        symbols_data = set(symbols_index_data)

        hist_minus_data = symbols_hist - symbols_data
        data_minus_hist = symbols_data - symbols_hist

        if len(hist_minus_data) > 0:
            print('HistMore Hist + [{}] , Data - [{}], Diff: {}'.format(
                len(symbols_index_hist), len(symbols_index_data), hist_minus_data))
        if len(data_minus_hist) > 0:
            print('DataMore Hist - [{}], Data + [{}], Diff: {}'.format(
                len(symbols_index_hist), len(symbols_index_data), data_minus_hist))

    def test_inc_exc_index_data(self, indices='default', end_year='2100'):
        """
        Test whether yearly index files are in line with tblHistIndex
        :param indices:
        :param end_year:
        :return:
        """

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )

        qry = 'SELECT * FROM tblHistIndex WHERE IndexName in {} AND Date <= "{}1231" ' \
              'ORDER BY IndexName ASC, Date ASC'.format(str(indices), end_year)
        index_hist_all = pd.read_sql(qry, self.conn)

        years = [year for year in self.YEARS if int(year) <= int(end_year)]

        year_data = dict()
        for year in years:
            year_data[year] = pd.read_csv('{}{}.csv'.format(self.INC_EXC_PATH, year))

        for index in indices:
            print('Processing {}...'.format(index))
            index_hist = index_hist_all[index_hist_all.IndexName == index]
            prev_date = '19000101'

            for date in index_hist['Date'].unique():
                prev_yr, curr_yr = prev_date[0:4], date[0:4]
                if prev_yr != curr_yr:
                    if prev_yr != '1900':
                        print('Year end check')
                        self.compare_index_data(index, prev_yr, prev_date, date, index_hist, index_year_data)
                    index_year_data = year_data[curr_yr][year_data[curr_yr][index] == 1]

                self.compare_index_data(index, curr_yr, prev_date, date, index_hist, index_year_data)

                prev_date = date

    def create_amibroker_import_files(self, path, start_date='19000101', type='delta'):

        start_year = start_date[0:4]
        years = [year for year in self.YEARS if year >= start_year]

        for year in years:
            print('creating files for year', year)
            if year == start_year:
                qry = """SELECT Symbol, Date, Open, High, Low, Close, Volume 
                           FROM tblModDump{} WHERE Date >= '{}'
                          ORDER BY Symbol ASC, Date ASC""".format(year, start_date)
                qry_mod = """SELECT Symbol, Date, AdjustedOpen Open, AdjustedHigh High, AdjustedLow Low, 
                                    AdjustedClose Close, Volume 
                               FROM tblModDump{0} WHERE Date >= '{1}' AND AdjustedOpen IS NOT NULL 
                              UNION
                             SELECT Symbol, Date, Open, High, Low, Close, Volume 
                               FROM tblModDump{0} WHERE Date >= '{1}' AND AdjustedOpen IS  NULL""".format(
                    year, start_date)
            else:
                qry = """SELECT Symbol, Date, Open, High, Low, Close, Volume 
                           FROM tblModDump{}
                          ORDER BY Symbol ASC, Date ASC""".format(year)
                qry_mod = """SELECT Symbol, Date, AdjustedOpen Open, AdjustedHigh High, AdjustedLow Low, 
                                    AdjustedClose Close, Volume
                               FROM tblModDump{0} WHERE AdjustedOpen IS NOT NULL 
                              UNION
                             SELECT Symbol, Date, Open, High, Low, Close, Volume 
                               FROM tblModDump{0} WHERE AdjustedOpen IS  NULL""".format(year)

            name = start_date if type == 'delta' else year
            df = pd.read_sql(qry, self.conn)
            df.to_csv('{}{}.csv'.format(path, name), sep=',', index=False)
            df = pd.read_sql(qry_mod, self.conn)
            df.sort_values(['Symbol', 'Date'], ascending=[True, True])
            df.Symbol = df.Symbol + '.A'
            df.to_csv('{}A.{}.csv'.format(path, name), sep=',', index=False)

    def create_amibroker_import_files_index_incexc(self, index, path, start_date='19000101', type='delta'):

        start_year = start_date[0:4]

        if type == 'full':
            years = [year for year in self.YEARS if year >= start_year]

            for year in years:
                print('creating files for year', year)
                df = pd.read_csv('{}{}.csv'.format(self.INC_EXC_PATH, year))
                df = df[df.Date >= int(start_date)]
                df.Symbol = df.Symbol + '.{}'.format(self.NIFTY_INDICES_AMI[index])
                df['Open'], df['High'], df['Low'], df['Close'] = 0, 0, 0, 1
                df = df[['Symbol', 'Date', 'Open', 'High', 'Low', 'Close', index]]
                df.to_csv('{}{}.{}.csv'.format(path, self.NIFTY_INDICES_AMI[index], year), sep=',', index=False)

        elif type == 'delta':  # Works only if start_date >= max date in tblHistIndex
            last_idx_change_date = pd.read_sql_query('SELECT MAX(Date) MaxDate FROM tblHistIndex', self.conn)
            if start_date < last_idx_change_date.MaxDate[0]:
                print('start date {} less than last index change date {}, quitting....')
            else:
                print('creating file for start date', start_date)
                df_idx = pd.read_csv(self.INDEX_COMPONENTS_CURR)
                df_idx = df_idx[df_idx.Index == index]
                df_idx = df_idx.set_index('Symbol')
                df_idx[index] = 1

                df = pd.read_csv('{}{}.csv'.format(self.INC_EXC_PATH, start_date))
                df = df[df.Date >= int(start_date)]
                df = df.set_index('Symbol')
                df = df.assign(index=df_idx[index]).fillna(0)
                df['Symbol'] = df.index
                df['Open'], df['High'], df['Low'], df['Close'] = 0, 0, 0, 1
                df = df[['Symbol', 'Date', 'Open', 'High', 'Low', 'Close', index]]
                df.Symbol = df.Symbol + '.{}'.format(self.NIFTY_INDICES_AMI[index])
                df.to_csv('{}{}.{}.csv'.format(path, self.NIFTY_INDICES_AMI[index], start_date), sep=',', index=False)
