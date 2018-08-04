"""
Created on Jun 10, 2018
@author: Souvik
@Program Function: Adjust NSE EOD Data
"""

import os
import dates, utils
import pandas as pd
import numpy as np
import sqlite3
from sqlalchemy import create_engine
from collections import OrderedDict
import openpyxl as xl


class DataDB:
    """ Historical Bhavcopy data"""

    # constants

    YEARS = ['1995', '1996', '1997', '1998', '1999', '2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007',
             '2008', '2009', '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018']

    MOD_PATH = 'nse_eod_modifiers/'
    BONUS_SPLITS_FILE = '{}bonus_splits.csv'.format(MOD_PATH)
    SYMBOL_CHANGE_FILE = '{}symbol_change.csv'.format(MOD_PATH)
    MULTIPLIERS_FILE = '{}multipliers.csv'.format(MOD_PATH)
    MULTIPLIER_SKIPS_FILE = '{}multipliers_skips.csv'.format(MOD_PATH)
    INDEX_CHANGE_DUMP_XL = '{}IndexInclExcl.xlsx'.format(MOD_PATH)
    INDEX_CHANGE_DUMP_CSV = '{}IndexInclExcl.csv'.format(MOD_PATH)
    INDEX_CHANGE_MOD_CSV = '{}IndexInclExclMod.csv'.format(MOD_PATH)
    INDEX_CHANGE_MOD_TEST_CSV = '{}IndexInclExclModTest.csv'.format(MOD_PATH)
    INDEX_COMPONENTS_CURR = '{}index_components_curr.csv'.format(MOD_PATH)
    INDEX_CHANGE_CSV = '{}index_inc_exc.csv'.format(MOD_PATH)
    SYMBOL_CHANGE_DUPLICATES_FILE = '{}symbol_change_duplicates.csv'.format(MOD_PATH)
    UNVERIFIED_SKIPPED_RECORDS_FILE = '{}unverified_skipped_records.csv'.format(MOD_PATH)
    UNVERIFIED_SELECTED_RECORDS_FILE = '{}unverified_selected_records.csv'.format(MOD_PATH)
    REPLACEBLE_SELECTED_RECORDS = '{}replaceble_selected_records.csv'.format(MOD_PATH)
    REPLACEBLE_SKIPPED_RECORDS = '{}replaceble_skipped_records.csv'.format(MOD_PATH)

    NIFTY_INDICES = {
        'NIFTY LargeMidcap 250': 'NIFTY_LARGEMIDCAP_250',
        'NIFTY Midcap 100': 'NIFTY_MIDCAP_100',
        'NIFTY Smallcap 100': 'NIFTY_SMALLCAP_100',
        'Nifty 100': 'NIFTY_100',
        'Nifty 200': 'NIFTY_200',
        'Nifty 50': 'NIFTY_50',
        'Nifty 500': 'NIFTY_500',
        'Nifty Alpha 50': 'NIFTY_ALPHA_50',
        'Nifty Auto': 'NIFTY_AUTO',
        'Nifty Bank': 'NIFTY_BANK',
        'Nifty Commodities': 'NIFTY_COMMODITIES',
        'Nifty Dividend Opportunities 50': 'NIFTY_DIVIDEND_OPPORTUNITIES_50',
        'Nifty Energy': 'NIFTY_ENERGY',
        'Nifty FMCG': 'NIFTY_FMCG',
        'Nifty Financial Services': 'NIFTY_FINANCIAL_SERVICES',
        'Nifty Growth Sectors 15': 'NIFTY_GROWTH_SECTORS_15',
        'Nifty High Beta 50': 'NIFTY_HIGH_BETA_50',
        'Nifty IT': 'NIFTY_IT',
        'Nifty India Consumption': 'NIFTY_INDIA_CONSUMPTION',
        'Nifty Infrastructure': 'NIFTY_INFRASTRUCTURE',
        'Nifty Low Volatility 50': 'NIFTY_LOW_VOLATILITY_50',
        'Nifty MNC': 'NIFTY_MNC',
        'Nifty Media': 'NIFTY_MEDIA',
        'Nifty Metal': 'NIFTY_METAL',
        'Nifty Midcap 50': 'NIFTY_MIDCAP_50',
        'Nifty Midcap Liquid 15': 'NIFTY_MIDCAP_LIQUID_15',
        'Nifty Next 50': 'NIFTY_NEXT_50',
        'Nifty PSE': 'NIFTY_PSE',
        'Nifty PSU Bank': 'NIFTY_PSU_BANK',
        'Nifty Pharma': 'NIFTY_PHARMA',
        'Nifty Realty': 'NIFTY_REALTY',
        'Nifty Services Sector': 'NIFTY_SERVICES_SECTOR',
        'Nifty100 Liquid 15': 'NIFTY100_LIQUID_15',
        'Nifty50 Value 20': 'NIFTY50_VALUE_20',
    }


    def __init__(self, db, type='EQ'):

        # variables

        self.TYPE = type

        print('Opening Bhavcopy database {}...'.format(db))
        self.conn = sqlite3.connect(db)
        self.engine = create_engine('sqlite:///{}'.format(db))

    def __del__(self):

        print('Closing DB connection..')
        self.conn.close()

    def truncate_table(self, table, msg=False):

        if msg is True:
            print('Truncating table {}'.format(table))

        c = self.conn.cursor()
        c.execute('''DELETE FROM {}'''.format(table))
        self.conn.commit()
        c.close()

    def load_multipliers(self, type='append'):

        multipliers = pd.read_csv(self.MULTIPLIERS_FILE)
        read_count = len(multipliers.index)
        print('Initiating loading of {} records into tblMultipliers'.format(read_count))

        if type == 'append':
            curr_multipliers = pd.read_sql_query('SELECT Symbol, Date, Multiplier FROM tblMultipliers', self.conn)
            print('{} records currently in tblMultipliers'.format(len(curr_multipliers.index)))
            multipliers = pd.concat([curr_multipliers, multipliers], axis=0)

        self.truncate_table('tblMultipliers', True)

        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        for symbol in symbol_changes.keys():
            multipliers['Symbol'] = np.where(multipliers.Symbol == symbol, symbol_changes[symbol], multipliers.Symbol)

        symbol_range_records = pd.read_sql('SELECT Symbol, StartDate FROM tblSymbolRange', self.conn)
        symbol_start_dates = dict(zip(symbol_range_records.Symbol, symbol_range_records.StartDate))
        #print(symbol_start_dates)

        c = self.conn.cursor()
        df_skips = pd.DataFrame()
        for symbol in multipliers['Symbol'].unique():
            print('processing', symbol)
            if symbol in symbol_start_dates:
                df = multipliers[(multipliers.Symbol == symbol) &
                             (multipliers.Date >= int(symbol_start_dates[symbol]))].copy()
                df['ResultantMultiplier'] = df['Multiplier'].cumprod()
                insert_rows = df.values.tolist()
                c.executemany('INSERT INTO tblMultipliers VALUES (?, ?, ?, ?)', insert_rows)
                self.conn.commit()
                df_skip = multipliers[(multipliers.Symbol == symbol) &
                                  (multipliers.Date < int(symbol_start_dates[symbol]))].copy()
                df_skips = pd.concat([df_skips, df_skip], axis=0)
            else:
                df_skip = multipliers[multipliers.Symbol == symbol].copy()
                df_skips = pd.concat([df_skips, df_skip], axis=0)

        c.close()

        df_skips.to_csv(self.MULTIPLIER_SKIPS_FILE, sep=',', index=False)


    def insert_records(self, df, table_name):
        """
        Insert passed records into table_name
        """

        df.to_sql(table_name, self.engine, index=False, if_exists='append')

    def load_table_from_csv(self, csv_path, start_date='1900-01-01', table_name='tblDump', type='append'):
        """
        Load table with historical data : Deprecated
        :param start_date: start date from which data files need to be considered
        :param csv_path: path of historical data in csv format
        :param table_name: table to be loaded
        :param type: 'append' if only new dates are to be loaded, 'refresh' if table needs to be refreshed
        :return: None
        """

        c = self.conn.cursor()

        if type == 'refresh':
            self.truncate_table(table_name, True)

        csv_files = [f for f in os.listdir(csv_path) if f.endswith('.txt') and f[0:10] >= start_date]
        csv_files.sort()

        print('Initiating loading of {} files'.format(len(csv_files)))

        col_names = ['Symbol', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume']
        year = '1900'
        skipped_records = pd.DataFrame(col_names)
        read_count, write_count, skipped_count = 0, 0, 0
        for file in csv_files:
            if year != file[0:4]:
                year = file[0:4]
                print('reading records for year {}'.format(year))
            df = pd.read_csv(csv_path + file, names=col_names, header=None)
            read_count = read_count + len(df.index)

            for idx, row in df.iterrows():
                insert_row = (row['Symbol'], row['Date'], row['Open'], row['High'], row['Low'], row['Close'],
                              row['Volume'])
                try:
                    c.execute('''INSERT INTO {} VALUES (?,?,?,?,?,?,?)'''.format(table_name), insert_row)

                except sqlite3.IntegrityError:
                    skipped_count = skipped_count + 1
                    skipped_record = [('Symbol', row['Symbol']), ('Date', row['Date']), ('Open', row['Open']),
                                      ('High', row['High']), ('Low', row['Low']), ('Close', row['Close']),
                                      ('Volume', row['Volume'])]
                    skipped_records = pd.concat([skipped_records, pd.DataFrame([OrderedDict(skipped_record)])], axis=0)

            if read_count - write_count > 20000:
                self.conn.commit()
                write_count = read_count - skipped_count
                print('inserted {} records, skipped {} records till now'.format(write_count, skipped_count))

        self.conn.commit()
        write_count = read_count - skipped_count
        c.close()
        skipped_records.to_csv(self.SKIPPED_RECORDS_FILE, sep=',', index=False)

        print('{} files processed, {} records inserted'.format(len(csv_files), write_count))

    def load_tbldumps_from_csv(self, csv_path, start_year='1995'):
        """
        Load table with historical data
        :param start_year: start year from which data files need to be considered
        :param csv_path: path of historical data in csv format
        :return: None
        """

        years_to_load = [year for year in self.YEARS if year >= start_year]

        c = self.conn.cursor()

        print('Truncating tables')
        for year in years_to_load:
            self.truncate_table('tblDump{}'.format(year))

        csv_files = [f for f in os.listdir(csv_path) if f.endswith('.txt') and f[0:4] >= start_year]
        csv_files.sort()

        print('Initiating loading of {} files'.format(len(csv_files)))

        col_names = ['Symbol', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume']
        year = '1900'
        skipped_records = pd.DataFrame()
        duplicate_records = pd.DataFrame()
        read_count, write_count, skipped_count = 0, 0, 0
        for file in csv_files:
            if year != file[0:4]:
                if year != '1900': # Write records to yearly tblDumps
                    year_unique = year_records.drop_duplicates(['Symbol', 'Date'], keep='first')
                    year_duplicate = year_records[year_records.duplicated(['Symbol', 'Date'], keep=False)]
                    year_skipped = year_records[year_records.duplicated(['Symbol', 'Date'], keep='first')]
                    self.insert_records(year_unique, table_name='tblDump{}'.format(year))
                    write_count = write_count + len(year_unique.index)
                    skipped_records = pd.concat([skipped_records, year_skipped], axis=0)
                    duplicate_records = pd.concat([duplicate_records, year_duplicate], axis=0)
                    print('{}: inserted {} records, duplicate {} records, skipped {} records'.format(
                        year, len(year_unique.index), len(year_duplicate.index), len(year_skipped.index)))

                year = file[0:4]
                print('reading records for year {}'.format(year))
                year_records = pd.DataFrame()

            df = pd.read_csv(csv_path + file, names=col_names, header=None)
            year_records = pd.concat([year_records, df], axis=0)
            read_count = read_count + len(df.index)

        # Write records for final year
        year_unique = year_records.drop_duplicates(['Symbol', 'Date'], keep='first')
        year_duplicate = year_records[year_records.duplicated(['Symbol', 'Date'], keep=False)]
        year_skipped = year_records[year_records.duplicated(['Symbol', 'Date'], keep='first')]
        self.insert_records(year_unique, table_name='tblDump{}'.format(year))
        write_count = write_count + len(year_unique.index)
        skipped_records = pd.concat([skipped_records, year_skipped], axis=0)
        duplicate_records = pd.concat([duplicate_records, year_duplicate], axis=0)
        print('{}: inserted {} records, duplicate {} records, skipped {} records'.format(year, len(year_unique.index),
                                                                                         len(year_skipped.index),
                                                                                         len(year_duplicate.index)))

        c.close()
        self.insert_records(skipped_records, table_name='tblSkipped')
        self.insert_records(duplicate_records, table_name='tblDuplicates')

        print('{} files processed, {} records read, {} records inserted, {} records duplicate, {} records skipped'.format(
            len(csv_files), read_count, write_count, len(duplicate_records.index), len(skipped_records.index)))

    def fetch_records(self, tables, symbols, start_date='19000101', end_date='21001231'):
        """
        Fetch records from tblDumps or tblModDumps for given parameters
        :param tables: 'tblDump' or 'tblModDump'
        :param symbols: [symbol1, symbol2, ...]
        :param start_date: 'YYYYMMDD'
        :param end_date: 'YYYYMMDD'
        :return: pandas dataframe with selected records
        """

        years = [year for year in self.YEARS if start_date[0:4] <= year <= end_date[0:4]]
        symbols_tuple = "('{}')".format(symbols[0]) if len(symbols) == 1 else str(tuple(symbols))

        if len(years) == 1:
            qry = """SELECT * FROM {}{} WHERE Symbol IN {} AND Date BETWEEN '{}' AND '{}'
                      ORDER BY Symbol ASC, Date ASC""".format(tables, years[0], symbols_tuple, start_date, end_date)
        else:
            qry_strings = []
            for i in range(0, len(years)):
                if i == 0:
                    qry_str = "SELECT * FROM {}{} WHERE Symbol IN {} AND Date >= '{}'".format(tables, years[i],
                                                                                              symbols_tuple, start_date)
                elif i == len(years) - 1:
                    qry_str = "SELECT * FROM {}{} WHERE Symbol IN {} AND Date <= '{}'".format(tables, years[i],
                                                                                              symbols_tuple, end_date)
                else:
                    qry_str = 'SELECT * FROM {}{} WHERE Symbol IN {}'.format(tables, years[i], symbols_tuple)

                qry_strings.append(qry_str)

            qry = '{} ORDER BY Symbol ASC, Date ASC'.format(' UNION '.join(qry_strings))

        return pd.read_sql_query(qry, self.conn)

    def int(self, val):
        try:
            return int(val)
        except ValueError:
            return 0

    def error(self, record, prev_record, next_record):

        p_open, n_open, r_open = self.int(prev_record['Open'].iloc[0]), self.int(next_record['Open'].iloc[0]), \
                                 self.int(record['Open'].iloc[0])
        p_high, n_high, r_high = self.int(prev_record['High'].iloc[0]), self.int(next_record['High'].iloc[0]), \
                                 self.int(record['High'].iloc[0])
        p_low, n_low, r_low = self.int(prev_record['Low'].iloc[0]), self.int(next_record['Low'].iloc[0]), \
                                 self.int(record['Low'].iloc[0])
        p_close, n_close, r_close = self.int(prev_record['Close'].iloc[0]), self.int(next_record['Close'].iloc[0]), \
                              self.int(record['Close'].iloc[0])
        p_volume, n_volume, r_volume = self.int(prev_record['Volume'].iloc[0]), self.int(next_record['Volume'].iloc[0]), \
                                    self.int(record['Volume'].iloc[0])

        open_e = (p_open + n_open) / 2 - r_open
        high_e = (p_high + n_high) / 2 - r_high
        low_e = (p_low + n_low) / 2 - r_low
        close_e = (p_close + n_close) / 2 - r_close
        volume_e = (p_volume + n_volume) / 2 - r_volume

        return {'price_error': open_e ** 2 + high_e ** 2 + low_e ** 2 + close_e ** 2,
                'volume_error': volume_e ** 2}
        
    def calculate_skipped_record_errors(self):

        skipped_qry = 'SELECT DISTINCT * from tblSkipped'
        skipped_records = pd.read_sql_query(skipped_qry, self.conn)

        unverified_skipped_records = pd.DataFrame()
        unverified_selected_records = pd.DataFrame()
        replaceable_selected_records = pd.DataFrame()
        replaceable_skipped_records = pd.DataFrame()

        for symbol in skipped_records['Symbol'].unique():
            print('checking for {}'.format(symbol))
            symbol_records = skipped_records[skipped_records.Symbol == symbol]
            for date in symbol_records['Date'].unique():
                date_records = symbol_records[symbol_records.Date == date]
                prev_sel_record_qry = '''SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date < "{}"
                                          ORDER BY Date DESC LIMIT 1'''.format(date[0:4], symbol, date)
                sel_record_qry = '''SELECT * FROM tblDump{} 
                                     WHERE Symbol = "{}" AND Date = {}'''.format(date[0:4], symbol, date)
                next_sel_record_qry = '''SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date > {}
                                          ORDER BY Date ASC LIMIT 1'''.format(date[0:4], symbol, date)

                prev_sel_record = pd.read_sql_query(prev_sel_record_qry, self.conn)
                sel_record = pd.read_sql_query(sel_record_qry, self.conn)
                next_sel_record = pd.read_sql_query(next_sel_record_qry, self.conn)

                if len(prev_sel_record.index) == 0 and len(next_sel_record.index) == 0:
                    for i in range(0, len(date_records.index)):
                        unverified_skipped_records = pd.concat([unverified_skipped_records,
                                                                date_records.iloc[i:i + 1, :]], axis=0)
                        unverified_selected_records = pd.concat([unverified_selected_records, sel_record], axis=0)
                else:
                    if len(prev_sel_record.index) == 0 and len(next_sel_record.index) > 0:
                        prev_sel_record = next_sel_record
                    elif len(prev_sel_record.index) > 0 and len(next_sel_record.index) == 0:
                        next_sel_record = prev_sel_record

                    sel_error = self.error(sel_record, prev_sel_record, next_sel_record)
                    skipped_error = []
                    for i in range(0, len(date_records.index)):
                        skipped_error.append(self.error(date_records.iloc[i:i+1, :], prev_sel_record, next_sel_record))
                        if skipped_error[i]['price_error'] < sel_error['price_error'] or \
                                skipped_error[i]['volume_error'] < sel_error['volume_error']:
                            replaceable_skipped_records = pd.concat([replaceable_skipped_records,
                                                                    date_records.iloc[i:i + 1, :]], axis=0)
                            replaceable_selected_records = pd.concat([replaceable_selected_records, sel_record], axis=0)

        unverified_skipped_records.to_csv(self.UNVERIFIED_SKIPPED_RECORDS_FILE, sep=',', index=False)
        unverified_selected_records.to_csv(self.UNVERIFIED_SELECTED_RECORDS_FILE, sep=',', index=False)
        replaceable_selected_records.to_csv(self.REPLACEBLE_SELECTED_RECORDS, sep=',', index=False)
        replaceable_skipped_records.to_csv(self.REPLACEBLE_SKIPPED_RECORDS, sep=',', index=False)

    def load_dump_replace_records(self):

        self.truncate_table('tblDumpReplace', True)

        c = self.conn.cursor()

        replaceable_selected_records = pd.read_csv(self.REPLACEBLE_SELECTED_RECORDS)
        replaceable_skipped_records = pd.read_csv(self.REPLACEBLE_SKIPPED_RECORDS)

        for i in range(0, len(replaceable_selected_records.index)):
            sel_record = replaceable_selected_records.iloc[i:i+1, :]
            skipped_record = replaceable_skipped_records.iloc[i:i + 1, :]
            #print(skipped_record)
            insert_row = (skipped_record['Symbol'].iloc[0], str(skipped_record['Date'].iloc[0]),
                          skipped_record['Open'].iloc[0], skipped_record['High'].iloc[0], skipped_record['Low'].iloc[0],
                          skipped_record['Close'].iloc[0], int(skipped_record['Volume'].iloc[0]))
            #print(insert_row)

            if sel_record['Symbol'].iloc[0] == skipped_record['Symbol'].iloc[0] and \
                sel_record['Date'].iloc[0] == skipped_record['Date'].iloc[0]:
                try:
                    c.execute('''INSERT INTO tblDumpReplace VALUES (?,?,?,?,?,?,?)''', insert_row)
                    print(insert_row, "inserted")
                except sqlite3.IntegrityError:
                    print(insert_row, "insert failed")
            else:
                print(insert_row, "skipped")

        self.conn.commit()
        c.close()

    def save_symbols_date_range(self, append_from='1995'):
        """
        Save all symbols start and end dates in csv file
        :param append_from: start year for appending
        :return:
        """

        exist_symbols_range = pd.read_sql_query('''SELECT * FROM tblSymbolRange''', self.conn)
        self.truncate_table('tblSymbolRange', True)

        years = [year for year in self.YEARS if year >= append_from]

        qry_strings_dump, qry_strings_mod = [], []
        for year in years:
            qry_strings_dump.append("SELECT Symbol, 'Dump' TableSource, MIN(Date) MinDate, MAX(Date) MaxDate "
                                    "FROM tblDump{} GROUP BY Symbol".format(year))
            qry_strings_mod.append("SELECT Symbol, 'ModDump' TableSource, MIN(Date) MinDate, MAX(Date) MaxDate "
                                   "FROM tblModDump{} GROUP BY Symbol".format(year))

        qry_dump = """SELECT Symbol, 'Dump' TableSource, MIN(MinDate) StartDate, MAX(MaxDate) EndDate FROM ({})
         GROUP BY Symbol""".format(' UNION '.join(qry_strings_dump))
        new_symbols_range_dump = pd.read_sql_query(qry_dump, self.conn)

        qry_mod = """SELECT Symbol, 'ModDump' TableSource, MIN(MinDate) StartDate, MAX(MaxDate) EndDate FROM ({})
         GROUP BY Symbol""".format(' UNION '.join(qry_strings_mod))
        new_symbols_range_mod = pd.read_sql_query(qry_mod, self.conn)

        new_symbols_range = pd.concat([new_symbols_range_dump, new_symbols_range_mod], axis=0)
        symbols_range = pd.concat([exist_symbols_range, new_symbols_range], axis=0)
        self.insert_records(symbols_range, 'tblSymbolRange')

        if len(new_symbols_range.index) > 0:
            print('Adjusting for existing records')
            qry = 'SELECT Symbol, TableSource, MIN(StartDate) StartDate, MAX(EndDate) EndDate FROM tblSymbolRange ' \
                  'GROUP BY Symbol, TableSource'
            final_recs = pd.read_sql_query(qry, self.conn)
            self.truncate_table('tblSymbolRange', True)
            self.insert_records(final_recs, 'tblSymbolRange')

    def load_modified_tbldumps(self, start_year='1995', end_year='2018'):
        """
        Load tblModDumps from tblDumps with symbol changes and corrected records
        :param start_year: start year from which tables need to be updated
        :return: None
        """

        symbol_change_records = pd.read_csv(self.SYMBOL_CHANGE_FILE)
        symbol_changes = dict(zip(symbol_change_records.Old, symbol_change_records.New))

        years_to_load = [year for year in self.YEARS if start_year <= year <= end_year]

        c = self.conn.cursor()

        print('Truncating tables')
        for year in years_to_load:
            self.truncate_table('tblModDump{}'.format(year))

        df_duplicate = pd.DataFrame()
        for year in years_to_load:
            print('updating for year', year)
            df = pd.read_sql_query('''SELECT * FROM tblDump{}'''.format(year), self.conn)
            df_replace = pd.read_sql_query('''SELECT * FROM tblDumpReplace WHERE Date LIKE "{}%"'''.format(year),
                                           self.conn)
            for symbol in symbol_changes.keys():
                df['Symbol'] = np.where(df.Symbol == symbol, symbol_changes[symbol], df.Symbol)
                df_replace['Symbol'] = np.where(df_replace.Symbol == symbol, symbol_changes[symbol], df_replace.Symbol)
            df = pd.concat([df_replace, df], axis=0)
            df_unique = df.drop_duplicates(['Symbol', 'Date'], keep='first').copy()
            df_unique['AdjustedOpen'] = np.nan
            df_unique['AdjustedHigh'] = np.nan
            df_unique['AdjustedLow'] = np.nan
            df_unique['AdjustedClose'] = np.nan
            self.insert_records(df_unique, 'tblModDump{}'.format(year))
            df_duplicate = pd.concat([df_duplicate, df[df.duplicated(['Symbol', 'Date'], keep='first')]], axis=0)

        if len(df_duplicate.index) > 0:
            df_duplicate.to_csv(self.SYMBOL_CHANGE_DUPLICATES_FILE, sep=',', index=False)

        c.close()

    def update_records(self, df, tables='tblModDump'):
        """
        Update passed records into table_name
        """

        c = self.conn.cursor()

        symbol = df.iloc[0]['Symbol']
        start_date = df.iloc[0]['Date']
        years = [year for year in self.YEARS if year >= start_date[0:4]]

        for year in years:
            if year == start_date[0:4]:
                delete_qry = '''DELETE FROM {}{} WHERE Symbol = "{}" AND Date >= "{}"'''.format(tables, year, symbol,
                                                                                                start_date)
            else:
                delete_qry = '''DELETE FROM {}{} WHERE Symbol = "{}"'''.format(tables, year, symbol)
            c.execute(delete_qry)
            df_year = df[(df.Date >= '{}0101'.format(year)) & (df.Date <= '{}1231'.format(year))]
            c.executemany('INSERT INTO {}{} VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'.format(tables, year),
                          df_year.values.tolist())
            self.conn.commit()



        #df.to_sql(table_name, self.engine, index=False, if_exists='append')

    def update_adjusted_price(self, start_date='19000101'):

        multipliers_qry = '''SELECT * FROM tblMultipliers WHERE Date >= "{}" ORDER BY Symbol, Date'''.format(start_date)
        multipliers = pd.read_sql_query(multipliers_qry, self.conn)

        #print(multipliers)
        #symbols_list = ["20MICRONS", "AARTIDRUGS", "AARTIIND"]

        c = self.conn.cursor()

        for symbol in multipliers['Symbol'].unique():
        #for symbol in symbols_list:
            symbol_multipliers = multipliers[multipliers.Symbol == symbol]
            symbol_records = self.fetch_records('tblModDump', [symbol], start_date=symbol_multipliers.iloc[0]['Date'])
            symbol_records_update = pd.DataFrame()
            for idx, row in symbol_multipliers.iterrows():
                #print(row)
                if len(symbol_records_update) > 0:
                    symbol_records_update = symbol_records_update[symbol_records_update.Date < row['Date']]
                temp_records = symbol_records[symbol_records.Date >= row['Date']].copy()
                temp_records['AdjustedOpen'] = temp_records['Open'] * row['ResultantMultiplier']
                temp_records['AdjustedHigh'] = temp_records['High'] * row['ResultantMultiplier']
                temp_records['AdjustedLow'] = temp_records['Low'] * row['ResultantMultiplier']
                temp_records['AdjustedClose'] = temp_records['Close'] * row['ResultantMultiplier']
                symbol_records_update = pd.concat([symbol_records_update, temp_records], axis=0)
            if len(symbol_records_update) > 0:
                first_date = symbol_records_update.iloc[0]['Date']
                delete_qry = '''DELETE FROM tblModDump{} WHERE Symbol = "{}" AND Date >= "{}"'''.format(first_date[0:4],
                                                                                                        symbol,
                                                                                                        first_date)
                c.execute(delete_qry)
                self.conn.commit()
                self.update_records(symbol_records_update)
                print('update complete for', symbol)
            else:
                print('update skipped for', symbol)

        c.close()

    def index_change_xl_to_csv(self):

        wb_idx = xl.load_workbook(self.INDEX_CHANGE_DUMP_XL)

        sheets = wb_idx.sheetnames

        #idx_records = pd.DataFrame('Index', 'Date', 'SymbolName', 'ChangeType')
        cols = ['Index', 'Date', 'SymbolName', 'ChangeType']
        records = []

        for sheet in sheets:
            sheet_name = wb_idx[sheet]
            for row in range(2, sheet_name.max_row + 1):
                records.append([sheet_name['A' + str(row)].value,
                                sheet_name['B' + str(row)].value,
                                sheet_name['C' + str(row)].value,
                                sheet_name['D' + str(row)].value])

        df = pd.DataFrame(records, columns=cols)
        df.to_csv(self.INDEX_CHANGE_DUMP_CSV, sep=',', index=False)

    def check_symbol_dates_old(self):

        df = pd.read_csv(self.INDEX_CHANGE_CSV)

        for idx, row in df.iterrows():
            qry = 'SELECT * FROM tblModDump{} WHERE Symbol = "{}" AND Date = "{}"'.format(str(row['Date'])[0:4],
                                                                                          row['Symbol'], row['Date'])
            record = pd.read_sql_query(qry, self.conn)

            if len(record.index) == 0:
                qry = 'SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date = "{}"'.format(str(row['Date'])[0:4],
                                                                                           row['Symbol'],
                                                                                           row['Date'])
                record = pd.read_sql_query(qry, self.conn)
                if len(record.index) > 0:
                    print(row['Symbol'], row['Date'], 'found in tblDump')
                else:
                    qry_prev = '''SELECT * FROM tblModDump{} WHERE Symbol = "{}" AND Date < "{}"
                                   LIMIT 1'''.format(str(row['Date'])[0:4], row['Symbol'], row['Date'])
                    record_prev = pd.read_sql_query(qry_prev, self.conn)
                    qry_next = '''SELECT * FROM tblModDump{} WHERE Symbol = "{}" AND Date > "{}"
                                                       LIMIT 1'''.format(str(row['Date'])[0:4], row['Symbol'],
                                                                         row['Date'])
                    record_next = pd.read_sql_query(qry_next, self.conn)
                    if len(record_prev.index) == 0 and len(record_next.index) == 0:
                        qry_prev = '''SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date < "{}"
                                                           LIMIT 1'''.format(str(row['Date'])[0:4], row['Symbol'],
                                                                             row['Date'])
                        record_prev = pd.read_sql_query(qry_prev, self.conn)
                        qry_next = '''SELECT * FROM tblDump{} WHERE Symbol = "{}" AND Date > "{}"
                                                                               LIMIT 1'''.format(str(row['Date'])[0:4],
                                                                                                 row['Symbol'],
                                                                                                 row['Date'])
                        record_next = pd.read_sql_query(qry_next, self.conn)
                        if len(record_prev.index) == 0 and len(record_next.index) == 0:
                            print(row['Symbol'], row['Date'], 'not found')
                        else:
                            if len(record_prev.index) > 0:
                                print(row['Symbol'], row['Date'], 'tblDump', 'prev', record_prev['Date'][0])
                            if len(record_next.index) > 0:
                                print(row['Symbol'], row['Date'], 'tblDump', 'next', record_next['Date'][0])
                    else:
                        if len(record_prev.index) > 0:
                            print(row['Symbol'], row['Date'], 'tblModDump', 'prev', record_prev['Date'][0])
                        if len(record_next.index) > 0:
                            print(row['Symbol'], row['Date'], 'tblModDump', 'next', record_next['Date'][0])

    def check_symbol_dates(self):

        df = pd.read_csv(self.INDEX_CHANGE_CSV)
        range_d = pd.read_csv('nse_eod_modifiers/symbol_range.csv')
        range_mod = pd.read_csv('nse_eod_modifiers/symbol_range_mod.csv')

        for idx, row in df.iterrows():
            s_range_mod = range_mod[range_mod.Symbol == row['Symbol']]
            s_range = range_d[range_d.Symbol == row['Symbol']]
            if not s_range_mod.empty:
                if s_range_mod.iloc[0]['StartDate'] < row['Date'] < s_range_mod.iloc[0]['EndDate']:
                    print('{},found in tblMod within range,{},{},{}'.format(row['Symbol'], row['Date'],
                                                                            s_range_mod.iloc[0]['StartDate'],
                                                                            s_range_mod.iloc[0]['EndDate']))
                else:
                    if not s_range.empty:
                        if s_range.iloc[0]['StartDate'] < row['Date'] < s_range.iloc[0]['EndDate']:
                            print('{},found in tblDump within range,{},{},{}'.format(row['Symbol'], row['Date'],
                                                                                     s_range.iloc[0]['StartDate'],
                                                                                     s_range.iloc[0]['EndDate']))
                        else:
                            print('{},found in tblMod outside range,{},{},{}'.format(row['Symbol'], row['Date'],
                                                                                     s_range_mod.iloc[0]['StartDate'],
                                                                                     s_range_mod.iloc[0]['EndDate']))
            else:
                if not s_range.empty:
                    if s_range.iloc[0]['StartDate'] < row['Date'] < s_range.iloc[0]['EndDate']:
                        print('{},found in tblDump within range,{},{},{}'.format(row['Symbol'], row['Date'],
                                                                                 s_range.iloc[0]['StartDate'],
                                                                                 s_range.iloc[0]['EndDate']))
                    else:
                        print('{},found in tblDump outside range,{},{},{}'.format(row['Symbol'], row['Date'],
                                                                                  s_range.iloc[0]['StartDate'],
                                                                                  s_range.iloc[0]['EndDate']))
                else:
                    print('{},{},not found'.format(row['Symbol'], row['Date']))

    def load_historical_index_components(self):

        c = self.conn.cursor()

        self.truncate_table('tblHistIndex', True)

        idx_change_hist = pd.read_csv(self.INDEX_CHANGE_MOD_CSV)
        idx_components_curr = pd.read_csv(self.INDEX_COMPONENTS_CURR)

        for index in idx_components_curr['Index'].unique():
        #for index in ['Nifty 500']:
            print('Loading for', index)
            index_hist = idx_change_hist[idx_change_hist.Index == index].copy()
            #print(index_hist)
            change_dates = list(index_hist['Date'].unique())
            change_dates.sort(reverse=True)
            #print(change_dates)
            curr_symbol_recs = idx_components_curr[idx_components_curr.Index == index].copy()

            first = True
            insert_symbols = set()
            for change_date in change_dates:
                #print(idx_change_hist)
                excluded_symbols = index_hist[(index_hist.Date == change_date) &
                                              (index_hist.ChangeType == 'E')]['Symbol'].tolist()
                included_symbols = index_hist[(index_hist.Date == change_date) &
                                              (index_hist.ChangeType == 'I')]['Symbol'].tolist()
                #print(change_date, '.......')
                #print('excluded symbols', 'SEARCHEMIN' if 'SEARCHEMIN' in excluded_symbols else 'Missing')
                #print('included symbols', 'SEARCHEMIN' if 'SEARCHEMIN' in included_symbols else 'Missing')
                #print('excluded #############################')
                #for symbol in excluded_symbols:
                #    print(symbol)
                #print('included #############################')
                #for symbol in included_symbols:
                #    print(symbol)

                if first:
                    curr_symbols = curr_symbol_recs['Symbol'].tolist()
                    curr_dates = [change_date] * len(curr_symbols)
                    curr_index = [index] * len(curr_symbols)
                    curr_df = pd.DataFrame(np.column_stack([curr_index, curr_symbols, curr_dates]),
                                           columns=['Index', 'Symbol', 'Date'])
                    insert_rows = curr_df.values.tolist()
                    #print('first inserted symbols', 'SEARCHEMIN' if 'SEARCHEMIN' in curr_symbols else 'Missing')
                    c.executemany('INSERT INTO tblHistIndex VALUES (?, ?, ?)', insert_rows)
                    insert_symbols = list(set(curr_symbols).union(set(excluded_symbols)) - set(included_symbols))
                    first = False
                else:
                    insert_symbols = list(set(insert_symbols).union(set(included_symbols)) - set(excluded_symbols))
                    insert_dates = [change_date] * len(insert_symbols)
                    insert_index = [index] * len(insert_symbols)
                    insert_df = pd.DataFrame(np.column_stack([insert_index, insert_symbols, insert_dates]),
                                             columns=['Index', 'Symbol', 'Date'])
                    insert_rows = insert_df.values.tolist()
                    #print('inserted symbols', 'SEARCHEMIN' if 'SEARCHEMIN' in insert_symbols else 'Missing')
                    c.executemany('INSERT INTO tblHistIndex VALUES (?, ?, ?)', insert_rows)
                    insert_symbols = list(set(insert_symbols).union(excluded_symbols) - set(included_symbols))

        self.conn.commit()
        c.close()

    def symbols_index_hist_files(self, indices='default', start_year='1995', end_year='2100'):

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )
            #indices = ('Nifty 50', 'Nifty 100')

        years = [year for year in self.YEARS if start_year <= year <= end_year]

        c = self.conn.cursor()

        #qry = 'SELECT * FROM tblHistIndex WHERE IndexName in {} AND Date >= "{}"' \
        #      'ORDER BY IndexName ASC, Date ASC'.format(str(indices), prev_yr_last_date)
        #index_hist = pd.read_sql(qry, self.conn)

        #index_hist_first = index_hist.drop_duplicates(['IndexName'], keep='first')
        #first_dates = dict(zip(index_hist_first['IndexName'], index_hist_first['Date']))

        print('Loading Index History...')
        qry = """SELECT t1. * 
                   FROM tblHistIndex t1
                   JOIN (SELECT IndexName, Min(Date) FirstDate
                           FROM (SELECT DISTINCT IndexName, Max(Date) Date
                                   FROM tblHistIndex
                                  WHERE Date < '{}0101'
                                  GROUP BY IndexName
                                  UNION
                                 SELECT DISTINCT IndexName, Min(Date) Date
                                   FROM tblHistIndex
                                  WHERE Date >= '{}0101'
                                  GROUP BY IndexName)
                          GROUP BY IndexName) t2
                  WHERE t1.IndexName = t2.IndexName
                    AND t1.Date >= t2.FirstDate
                    AND t2.IndexName in {}
                  ORDER BY IndexName ASC, Date ASC""".format(start_year, start_year, str(indices))

        index_hist = pd.read_sql(qry, self.conn)
        index_hist_first = index_hist.drop_duplicates(['IndexName'], keep='first')
        first_dates = dict(zip(index_hist_first['IndexName'], index_hist_first['Date']))
        prev_dates = first_dates

        #symbols_for_next_year = {}
        #first, prev_date = True, '19000101'
        #for year in ['1996', '1997', '1998']:


        #prev_dates = first_dates
        for year in years:
            print('Processing {}...'.format(year))

            qry_yr = 'SELECT Symbol, Date FROM tblModDump{}'.format(year)
            yr_records = pd.read_sql(qry_yr, self.conn)
            yr_records['Date'] = pd.to_numeric(yr_records['Date'])
            for index in index_hist['IndexName'].unique():
                #if index not in prev_dates:
                #    prev_dates[index] = '19000101'
                print('Processing {} {}...'.format(year, index))
                yr_records[index] = [0] * yr_records['Symbol'].size
                index_hist_curr = index_hist[(index_hist.IndexName == index) &
                                             (index_hist.Date >= prev_dates[index])].copy()
                #(index_hist.Date >= '{}0101'.format(year))].copy()
                #print(list(index_hist_curr['Date'].unique()))
                in_scope_dates = [d for d in list(index_hist_curr['Date'].unique()) if int(d[0:4]) == int(year)]
                #print(in_scope_dates)
                #for date in index_hist_curr['Date'].unique():
                for date in in_scope_dates:
                    #if int(date[0:4]) != int(year) and int(prev_dates[index][0:4]) != int(year):
                    #    continue
                    #if condition change change to use first_date
                    #if int(year) <= int(first_dates[index][0:4]):
                    #    continue
                    print('................ Processing {}... prev date {}'.format(date, prev_dates[index]))

                    index_hist_prev_date = index_hist_curr[index_hist_curr.Date == prev_dates[index]].copy()
                    for symbol in index_hist_prev_date['Symbol'].unique():
                        #print(symbol, prev_dates[index])
                        yr_records.loc[(yr_records.Symbol == symbol) & (yr_records.Date >= int(prev_dates[index])) &
                                       (yr_records.Date < int(date)), index] = 1
                    prev_dates[index] = date
                #if int(date[0:4]) != int(year) and int(prev_dates[index][0:4]) != int(year):
                if prev_dates[index][0:4] != year:
                    #print('Skipped {}...'.format(prev_dates[index]))
                    #prev_dates[index] = date
                    continue
                #if year <= first_dates[index][0:4]:
                #    continue
                print('................ Processing {}...'.format(prev_dates[index]))
                index_hist_prev_date = index_hist_curr[index_hist_curr.Date == prev_dates[index]].copy()
                for symbol in index_hist_prev_date['Symbol'].unique():
                    #print(symbol, prev_dates[index])
                    yr_records.loc[(yr_records.Symbol == symbol) &
                                   (yr_records.Date >= int(prev_dates[index])), index] = 1
                #prev_dates[index] = date

            yr_records.to_csv('{}{}.csv'.format(self.MOD_PATH, year), sep=',', index=False)

        c.close()

    def test_inc_exc_mod_list(self):

        idx_change_hist = pd.read_csv(self.INDEX_CHANGE_MOD_CSV)
        idx_components_curr = pd.read_csv(self.INDEX_COMPONENTS_CURR)

        idx_change_hist_last = idx_change_hist.drop_duplicates(['Index', 'Symbol'], keep='first')

        idx_change_hist_last.to_csv('{}IndexInclExclModTestLast.csv'.format(self.MOD_PATH), sep=',', index=False)

        errors = {}
        for index in idx_change_hist_last['Index'].unique():
            errors[index] = []
            hist_last_curr_index = idx_change_hist_last[idx_change_hist_last.Index == index]
            idx_components_curr_index = idx_components_curr[idx_components_curr.Index == index]
            for idx, row in hist_last_curr_index.iterrows():
                if row['ChangeType'] == 'I' and row['Symbol'] not in idx_components_curr_index['Symbol']:
                    errors[index].append(row['Symbol'])

        print(errors)

    def compare_index_data(self, index, year_file, prev_date, date, index_hist, index_year_data):

        print('{}........................ {} Processing {} - {}...'.format(index, year_file, prev_date, date))

        symbols_index_hist = index_hist[index_hist.Date == prev_date]['Symbol'].unique()
        symbols_index_data = index_year_data[(index_year_data.Date >= int(prev_date)) &
                                             (index_year_data.Date < int(date))]['Symbol'].unique()

        symbols_hist = set(symbols_index_hist)
        symbols_data = set(symbols_index_data)

        hist_minus_data = symbols_hist - symbols_data
        data_minus_hist = symbols_data - symbols_hist

        if len(hist_minus_data) > 0:
            print('HistMore Hist + [{}] , Data - [{}], Diff: {}'.format(
                len(symbols_index_hist), len(symbols_index_data), hist_minus_data))
        if len(data_minus_hist) > 0:
            print('DataMore Hist - [{}], Data + [{}], Diff: {}'.format(
                len(symbols_index_hist), len(symbols_index_data), data_minus_hist))

    def test_inc_exc_index_data(self, indices='default', end_year='2100'):

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )
            #indices = ('Nifty 50', 'Nifty 100')

        qry = 'SELECT * FROM tblHistIndex WHERE IndexName in {} AND Date <= "{}1231" ' \
              'ORDER BY IndexName ASC, Date ASC'.format(str(indices), end_year)
        index_hist_all = pd.read_sql(qry, self.conn)

        years = [year for year in self.YEARS if int(year) <= int(end_year)]

        year_data = dict()
        for year in years:
            year_data[year] = pd.read_csv('{}{}.csv'.format(self.MOD_PATH, year))

        for index in indices:
            print('Processing {}...'.format(index))
            index_hist = index_hist_all[index_hist_all.IndexName == index]
            prev_date = '19000101'

            for date in index_hist['Date'].unique():
                prev_yr, curr_yr = prev_date[0:4], date[0:4]
                if prev_yr != curr_yr:
                    if prev_yr != '1900':
                        print('Year end check')
                        self.compare_index_data(index, prev_yr, prev_date, date, index_hist, index_year_data)
                    index_year_data = year_data[curr_yr][year_data[curr_yr][index] == 1]

                self.compare_index_data(index, curr_yr, prev_date, date, index_hist, index_year_data)

                prev_date = date




    def test_inc_exc_index_data_old2(self, indices='default', end_year='2100'):

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )
            #indices = ('Nifty 50', 'Nifty 100')

            qry = 'SELECT * FROM tblHistIndex WHERE IndexName in {} AND Date <= "{}1231" ' \
                  'ORDER BY IndexName ASC, Date ASC'.format(str(indices), end_year)
            index_hist_all = pd.read_sql(qry, self.conn)

            for index in indices:
                print('Processing {}...'.format(index))
                index_hist = index_hist_all[index_hist_all.IndexName == index]
                prev_date = '19000101'

                for date in index_hist['Date'].unique():
                    if prev_date == '19000101':
                        prev_date = date
                        year_data = pd.read_csv('{}{}.csv'.format(self.MOD_PATH, date[0:4]))
                        index_year_data = year_data[year_data[index] == 1]
                        index_two_yrs_data = index_year_data
                        continue
                    else:
                        if prev_date[0:4] != date[0:4]:
                            index_prev_year_data = index_year_data
                            year_data = pd.read_csv('{}{}.csv'.format(self.MOD_PATH, date[0:4]))
                            index_year_data = year_data[year_data[index] == 1]
                            index_two_yrs_data = pd.concat([index_prev_year_data, index_year_data], axis=0)

                    print('{}........................Processing {} prev date {}...'.format(index, date, prev_date))
                    symbols_index_hist = index_hist[index_hist.Date == prev_date]['Symbol'].unique()
                    symbols_index_data = index_two_yrs_data[(index_two_yrs_data.Date >= int(prev_date)) &
                                                            (index_two_yrs_data.Date < int(date))]['Symbol'].unique()

                    symbols_hist = set(symbols_index_hist)
                    symbols_data = set(symbols_index_data)

                    hist_minus_data = symbols_hist - symbols_data
                    data_minus_hist = symbols_data - symbols_hist

                    if len(hist_minus_data) > 0:
                        print('HistMore Hist + [{}] , Data - [{}], Diff: {}'.format(
                            len(symbols_index_hist), len(symbols_index_data), hist_minus_data))
                    if len(data_minus_hist) > 0:
                        print('DataMore Hist - [{}], Data + [{}], Diff: {}'.format(
                            len(symbols_index_hist), len(symbols_index_data), data_minus_hist))

                    prev_date = date

    def test_inc_exc_index_data_old(self, indices='default', till='2100'):

        if indices == 'default':
            indices = (
                'NIFTY LargeMidcap 250',
                'NIFTY Midcap 100',
                'NIFTY Smallcap 100',
                'Nifty 100',
                'Nifty 200',
                'Nifty 50',
                'Nifty 500',
                'Nifty Alpha 50',
                'Nifty High Beta 50',
                'Nifty Low Volatility 50',
                'Nifty Midcap 50',
                'Nifty Midcap Liquid 15',
                'Nifty Next 50',
                'Nifty100 Liquid 15',
            )
            #indices = ('Nifty 50', 'Nifty 100')

            qry = 'SELECT * FROM tblHistIndex WHERE IndexName in {} AND Date <= "{}1231" ' \
                  'ORDER BY IndexName ASC, Date ASC'.format(str(indices), till)
            index_hist_all = pd.read_sql(qry, self.conn)

            for index in indices:
                print('Processing {}...'.format(index))
                index_hist = index_hist_all[index_hist_all.IndexName == index]
                prev_date = '19000101'

                for date in index_hist['Date'].unique():
                    if prev_date == '19000101':
                        prev_date = date
                        year_data = pd.read_csv('{}{}.csv'.format(self.MOD_PATH, date[0:4]))
                        index_year_data = year_data[year_data[index] == 1]
                        index_two_yrs_data = index_year_data
                        continue
                    else:
                        if prev_date[0:4] != date[0:4]:
                            index_prev_year_data = index_year_data
                            year_data = pd.read_csv('{}{}.csv'.format(self.MOD_PATH, date[0:4]))
                            index_year_data = year_data[year_data[index] == 1]
                            index_two_yrs_data = pd.concat([index_prev_year_data, index_year_data], axis=0)

                    print('{}........................Processing {} prev date {}...'.format(index, date, prev_date))
                    symbols_index_hist = index_hist[index_hist.Date == prev_date]['Symbol'].unique()
                    symbols_index_data = index_two_yrs_data[(index_two_yrs_data.Date >= int(prev_date)) &
                                                            (index_two_yrs_data.Date < int(date))]['Symbol'].unique()

                    symbols_hist = set(symbols_index_hist)
                    symbols_data = set(symbols_index_data)

                    hist_minus_data = symbols_hist - symbols_data
                    data_minus_hist = symbols_data - symbols_hist

                    if len(hist_minus_data) > 0:
                        print('HistMore Hist + [{}] , Data - [{}], Diff: {}'.format(
                            len(symbols_index_hist), len(symbols_index_data), hist_minus_data))
                    if len(data_minus_hist) > 0:
                        print('DataMore Hist - [{}], Data + [{}], Diff: {}'.format(
                            len(symbols_index_hist), len(symbols_index_data), data_minus_hist))

                    prev_date = date





























